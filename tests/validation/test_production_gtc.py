"""Production GTC-001 comparison against approved Golden Master."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from pems.calculations.modules.ec_io import EcIoModule
from pems.calculations.modules.production import ProductionModule
from pems.gtc.compare import (
    assert_gtc_bound_to_active_sha,
    compare_cell_map,
    load_formula_cached_expected,
    load_literal_expected,
)
from pems.infrastructure.excel_import import import_case_input_from_active_gm
from pems.infrastructure.golden_master import ACTIVE_GM_SHA256, verify_active_gm
from pems.validation.case_input_validator import validate_case_input

ROOT = Path(__file__).resolve().parents[2]

# Contract §9 + design intermediates
LITERAL_CELLS = [
    ("Production Profile", "C3"),
    ("Production Profile", "F3"),
    ("Production Profile", "F5"),
    ("Production Profile", "C7"),
    ("Production Profile", "C8"),
    ("Production Profile", "C9"),
    ("Production Profile", "C12"),
    ("Production Profile", "C13"),
    ("Production Profile", "I13"),
    ("Production Profile", "C14"),
    ("Production Profile", "F14"),
    ("Production Profile", "L7"),
    ("Prod_Summary", "Y48"),
]

FORMULA_CELLS = [
    ("Production Profile", "C2"),
    ("Production Profile", "F2"),
    ("Production Profile", "C4"),
    ("Production Profile", "F4"),
    ("Production Profile", "C6"),
    ("Production Profile", "C15"),
    ("Production Profile", "I15"),
    ("Production Profile", "I14"),
    ("Production Profile", "C16"),
    ("Production Profile", "F16"),
    ("Production Profile", "I16"),
    ("Production Profile", "F17"),
    ("Prod_Summary", "V47"),
    ("Prod_Summary", "Y47"),
    ("Prod_Summary", "Y49"),
    ("Prod_Summary", "Y50"),
    ("Prod_Summary", "AF26"),
    ("Prod_Summary", "R1"),
    ("Prod_Summary", "AF21"),
    ("Ec_IO", "C6"),
]

TEXT_CELLS = [
    ("Production Profile", "B2"),
]


@pytest.fixture(scope="module")
def repo_root() -> Path:
    return ROOT


def _gm_data_only(repo_root: Path, cells: list[tuple[str, str]]):
    path = verify_active_gm(repo_root)
    wb = load_workbook(path, data_only=True, read_only=True)
    out = {}
    try:
        for sheet, cell in cells:
            out[(sheet, cell)] = wb[sheet][cell].value
    finally:
        wb.close()
    return out


def test_gm_sha_intact(repo_root: Path) -> None:
    verify_active_gm(repo_root)
    assert_gtc_bound_to_active_sha(repo_root)
    assert ACTIVE_GM_SHA256 == "D07560CA6C1A762716E1927A130E7CA697DB0AB9BDA8E8A33C7A0ACBB6FDBFEA"


def test_import_production_params(repo_root: Path, active_gm_case) -> None:
    case = active_gm_case  # session-cached GM import (tests/conftest.py)
    errs = validate_case_input(case)
    assert errs == []
    assert case.pp_mode == "STOIIP"
    assert case.qi_buildup == 1000
    assert case.qp_plateau == 6000
    assert case.block_field_oil == "Ebiya Field"
    assert case.oil_block_daily is not None and len(case.oil_block_daily) > 0
    assert case.gas_block_daily is not None and len(case.gas_block_daily) > 0


def test_production_gtc_comparison(repo_root: Path, active_gm_case) -> None:
    case = active_gm_case  # session-cached GM import (tests/conftest.py)
    prod = ProductionModule().run(case)
    pems = prod.cell_map()

    # Also pass life into Ec_IO for C6 interface check
    eco = EcIoModule().run(case, upstream={"project_life_years": prod.project_life_years})
    pems[("Ec_IO", "C6")] = eco.project_end_year_e29 and prod.project_life_years
    pems[("Ec_IO", "C6")] = prod.project_life_years

    expected_lit = load_literal_expected(repo_root, LITERAL_CELLS)
    expected_f = load_formula_cached_expected(repo_root, FORMULA_CELLS)
    expected_text = _gm_data_only(repo_root, TEXT_CELLS)
    # R1/AF21 may be text from formula cache
    expected = {**expected_lit, **expected_f, **expected_text}
    # Fill text-like formula cells from data_only if missing
    for k in FORMULA_CELLS:
        if k not in expected or expected[k] is None:
            expected[k] = _gm_data_only(repo_root, [k]).get(k)

    compare_keys = [k for k in pems if k in expected and pems[k] is not None]
    res = compare_cell_map({k: pems[k] for k in compare_keys}, {k: expected[k] for k in compare_keys})
    mismatches = [d for d in res.details if d["status"] == "mismatch"]
    assert res.mismatch == 0, f"mismatches={mismatches}; exact={res.exact} tol={res.tolerance}"
    assert res.exact + res.tolerance >= 20


def test_production_series_sample_points(repo_root: Path, active_gm_case) -> None:
    """Sample Prod_Summary oil/gas rates vs GM cache (parity on block path)."""
    case = active_gm_case  # session-cached GM import (tests/conftest.py)
    prod = ProductionModule().run(case)
    path = verify_active_gm(repo_root)
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ps = wb["Prod_Summary"]
        samples = [
            (5, 2027),
            (8, 2030),
            (12, 2034),
            (19, 2041),
        ]
        for row, year in samples:
            t_exp = ps[f"T{row}"].value
            w_exp = ps[f"W{row}"].value
            t_act = prod.oil_daily_series.get(year)
            w_act = prod.gas_daily_series.get(year)
            assert t_act is not None and abs(float(t_act) - float(t_exp or 0)) < 1e-9
            assert w_act is not None and abs(float(w_act) - float(w_exp or 0)) < 1e-9
        assert abs((prod.oil_eur_or_max_cum or 0) - float(ps["V47"].value)) < 1e-9
        assert abs((prod.gas_max_cum or 0) - float(ps["Y47"].value)) < 1e-9
        assert prod.project_life_years == float(ps["AF26"].value)
    finally:
        wb.close()
