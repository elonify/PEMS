"""Excel → CaseInput import (openpyxl, read-only).

Maps GM cells per EC_IO_PARAMETER_CONTRACT and PRODUCTION_PROFILE_CONTRACT.
Does not write workbooks.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from pems.domain.case_input import CaseInput
from pems.infrastructure.golden_master import resolve_gm_path, verify_active_gm

# Block sheet field header row 1: daily rate columns (3-column groups).
# Exclude selector helper columns (CS/CT selected oil; CN/CO selected gas year/rate)
# which may display the active field name via formula and would overwrite real headers.
_BLOCK_FIELD_DAILY_COLS = [
    "B",
    "E",
    "H",
    "K",
    "N",
    "Q",
    "T",
    "W",
    "Z",
    "AC",
    "AF",
    "AI",
    "AL",
    "AO",
    "AR",
    "AU",
    "AX",
    "BA",
    "BD",
    "BG",
    "BJ",
    "BM",
    "BP",
    "BS",
    "BV",
    "BY",
    "CB",
    "CE",
    "CH",
    "CK",
]


def _as_float(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return float(v)
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _as_int_year(v: Any) -> int | None:
    f = _as_float(v)
    if f is None:
        return None
    return int(f)


def _as_str(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        return v
    return str(v)


def _header_map(ws: Any, daily_cols: list[str]) -> dict[str, str]:
    """Map field name → daily column letter (first match wins)."""
    out: dict[str, str] = {}
    for col in daily_cols:
        name = ws[f"{col}1"].value
        if name is None:
            continue
        if isinstance(name, str) and name.startswith("="):
            continue
        key = str(name).strip()
        if key and key not in out:
            out[key] = col
    return out


def _annual_col(daily_col: str) -> str:
    """Daily column + 1 letter offset (B→C, E→F, …)."""
    # openpyxl column index
    from openpyxl.utils import column_index_from_string

    idx = column_index_from_string(daily_col)
    return get_column_letter(idx + 1)


def _read_year_series(
    ws: Any,
    daily_col: str,
    *,
    row_start: int = 4,
    row_end: int = 60,
    annual_col: str | None = None,
) -> tuple[list[list[float]], list[list[float]]]:
    daily: list[list[float]] = []
    annual: list[list[float]] = []
    ac = annual_col or _annual_col(daily_col)
    for row in range(row_start, row_end + 1):
        y = _as_int_year(ws[f"A{row}"].value)
        if y is None:
            # year may be formula; data_only cache required
            continue
        d = _as_float(ws[f"{daily_col}{row}"].value)
        a = _as_float(ws[f"{ac}{row}"].value)
        if d is None and a is None:
            continue
        daily.append([float(y), float(d or 0.0)])
        annual.append([float(y), float(a or 0.0)])
    return daily, annual


def _read_cap_allow_category_series(
    ws: Any,
    *,
    year_col: str = "FE",
    value_col: str,
    row_start: int = 5,
    row_end: int = 46,
) -> list[list[float]]:
    """Read Cap_Allow FE-aligned category column as [year, $mm] pairs."""
    out: list[list[float]] = []
    for row in range(row_start, row_end + 1):
        y = _as_int_year(ws[f"{year_col}{row}"].value)
        if y is None:
            continue
        v = _as_float(ws[f"{value_col}{row}"].value)
        out.append([float(y), float(v or 0.0)])
    return out


def import_case_input_from_workbook(
    workbook_path: Path | None = None,
    *,
    repo_root: Path | None = None,
    require_active_gm_hash: bool = False,
    data_only: bool = True,
    include_production: bool = True,
    include_costs: bool = True,
) -> CaseInput:
    """Import CaseInput from Excel.

    Prefer ``data_only=True`` for literal values (uses Excel value cache).
    """
    if require_active_gm_hash:
        path = verify_active_gm(repo_root)
    elif workbook_path is not None:
        path = workbook_path
    else:
        path = resolve_gm_path(repo_root)

    wb = load_workbook(path, data_only=data_only, read_only=True)
    try:
        eco = wb["Ec_IO"]
        eq = wb["Equity Dash"]

        case = CaseInput(
            equity_share_company_1=_as_float(eq["C4"].value),
            project_equity_total=_as_float(eq["C6"].value) if eq["C6"].value is not None else 1.0,
            project_start_year=_as_int_year(eco["C5"].value),
            production_days_per_year=_as_float(eco["C7"].value),
            oil_price_usd_bbl=_as_float(eco["C12"].value),
            price_escalator=_as_float(eco["C14"].value),
            hurdle_rate=_as_float(eco["C15"].value),
            gas_price_usd_mscf=_as_float(eco["C17"].value),
            gas_flare_penalty_usd_mscf=_as_float(eco["C18"].value),
            dom_gas_fraction=_as_float(eco["C19"].value),
            duties_rate=_as_float(eco["C20"].value),
            vat_rate=_as_float(eco["C21"].value),
            asset_salvage_frac_of_retention=_as_float(eco["C22"].value),
            nag_crl=_as_float(eco["C23"].value),
            nag_ita=_as_float(eco["C24"].value),
            nag_min_tax_rate=_as_float(eco["C25"].value),
            nag_cpr=_as_float(eco["C26"].value),
            history_year=_as_int_year(eco["D28"].value),
            complete_year=_as_int_year(eco["D30"].value),
            project_life_years=_as_float(eco["C6"].value)
            if data_only
            and not (isinstance(eco["C6"].value, str) and str(eco["C6"].value).startswith("="))
            else None,
            asset_analysis_type=_as_str(eco["C4"].value),
            block_field_oil=_as_str(eco["G18"].value),
            terrain=_as_str(eco["G20"].value),
            gas_utilization=_as_str(eco["G21"].value),
            licence_lease_status=_as_str(eco["G22"].value),
            pfs_contract_type=_as_str(eco["G24"].value),
            country=_as_str(eco["G25"].value),
            fiscal_regime_label=_as_str(eco["G26"].value),
            source="excel_import",
            source_path=str(path),
        )
        gas_field = eco["G19"].value
        if data_only and not (isinstance(gas_field, str) and gas_field.startswith("=")):
            case.block_field_gas = _as_str(gas_field)
        else:
            case.block_field_gas = case.block_field_oil

        if include_production and "Production Profile" in wb.sheetnames:
            pp = wb["Production Profile"]
            case.pp_mode = _as_str(pp["B2"].value)
            case.stoiip_inplace = _as_float(pp["C2"].value)
            case.giip_inplace = _as_float(pp["F2"].value)
            case.oil_rf = _as_float(pp["C3"].value)
            case.gas_rf = _as_float(pp["F3"].value)
            case.gor_scf_bbl = _as_float(pp["F5"].value)
            case.prod_start_lag_years = _as_float(pp["C7"].value)
            case.year_end_anchor = _as_int_year(pp["C8"].value)
            case.pp_days_in_year = _as_float(pp["C9"].value)
            case.eff_decline_rate = _as_float(pp["L7"].value)
            case.qi_buildup = _as_float(pp["C12"].value)
            case.qp_plateau = _as_float(pp["C13"].value)
            case.qel_end = _as_float(pp["I13"].value)
            case.t1_buildup_yrs = _as_float(pp["C14"].value)
            case.t2_plateau_yrs = _as_float(pp["F14"].value)

            if "Prod_Summary" in wb.sheetnames:
                case.gas_boe_factor = _as_float(wb["Prod_Summary"]["Y48"].value)

            if "Analysis" in wb.sheetnames:
                an = wb["Analysis"]
                case.analysis_oil_scale = _as_float(an["N8"].value) or 0.0
                case.analysis_gas_scale = _as_float(an["N9"].value) or 0.0
            else:
                case.analysis_oil_scale = 0.0
                case.analysis_gas_scale = 0.0

            # Selected field block series (GTC parity path)
            if "Block_Oil Data" in wb.sheetnames and case.block_field_oil:
                bo = wb["Block_Oil Data"]
                oil_map = _header_map(bo, _BLOCK_FIELD_DAILY_COLS)
                col = oil_map.get(case.block_field_oil.strip())
                if col:
                    d, a = _read_year_series(bo, col)
                    case.oil_block_daily = d
                    case.oil_block_annual = a

            if "Block_Gas Data" in wb.sheetnames:
                bg = wb["Block_Gas Data"]
                gas_map = _header_map(bg, _BLOCK_FIELD_DAILY_COLS)
                gname = (case.block_field_gas or case.block_field_oil or "").strip()
                col = gas_map.get(gname)
                if col:
                    d, a = _read_year_series(bg, col)
                    case.gas_block_daily = d
                    case.gas_block_annual = a

        # G23 cost mode field (formula =G18 on GM when data_only)
        g23 = eco["G23"].value
        if data_only and not (isinstance(g23, str) and str(g23).startswith("=")):
            case.cost_mode_field = _as_str(g23) or case.block_field_oil
        else:
            case.cost_mode_field = case.block_field_oil

        if include_costs and "Cap_Allow" in wb.sheetnames:
            ca = wb["Cap_Allow"]
            # Selected consolidated Cap_Allow path (post field select + escalated OPEX)
            case.oil_tc_exploration = _read_cap_allow_category_series(ca, value_col="FF")
            case.oil_tc_capex_wells = _read_cap_allow_category_series(ca, value_col="FG")
            case.oil_tc_capex_facilities = _read_cap_allow_category_series(ca, value_col="FH")
            case.oil_tc_opex = _read_cap_allow_category_series(ca, value_col="FI")
            case.oil_sln_by_year = _read_cap_allow_category_series(ca, value_col="GX")
            case.oil_acq_allowance_by_year = _read_cap_allow_category_series(ca, value_col="HC")
            case.acquisition_cost = _as_float(ca["HB5"].value)
            rates: list[float] = []
            for row in range(5, 10):
                rv = _as_float(ca[f"FR{row}"].value)
                if rv is not None:
                    rates.append(rv)
            case.ca_rates = rates if rates else [0.2, 0.2, 0.2, 0.2, 0.19]

            if "Block_TC" in wb.sheetnames:
                case.opex_escalation_rate = _as_float(wb["Block_TC"]["FW3"].value) or 0.0
            else:
                case.opex_escalation_rate = 0.0

            if "Cap_Allow Gas" in wb.sheetnames:
                cag = wb["Cap_Allow Gas"]
                case.gas_tc_exploration = _read_cap_allow_category_series(cag, value_col="FF")
                case.gas_tc_capex_wells = _read_cap_allow_category_series(cag, value_col="FG")
                case.gas_tc_capex_facilities = _read_cap_allow_category_series(cag, value_col="FH")
                case.gas_tc_opex = _read_cap_allow_category_series(cag, value_col="FI")
                case.gas_sln_by_year = _read_cap_allow_category_series(cag, value_col="GX")
                case.gas_acq_allowance_by_year = _read_cap_allow_category_series(cag, value_col="HC")

        return case
    finally:
        wb.close()


def import_case_input_from_active_gm(repo_root: Path | None = None) -> CaseInput:
    """Import from approved Golden Master after SHA verification."""
    return import_case_input_from_workbook(
        None, repo_root=repo_root, require_active_gm_hash=True, data_only=True
    )
