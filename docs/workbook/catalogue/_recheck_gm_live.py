"""Read-only recheck: confirmed GM vs live workbook. Does not modify either file."""
from __future__ import annotations

from pathlib import Path
import hashlib
from datetime import datetime
import json
import zipfile

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula

base = Path(__file__).resolve().parents[1]
CONF = base / "Workbook_History" / "Econ_Model_PEMS_confirmed_2026-08-03.xlsx"
LIVE = base / "Econ_Model_PEMS.xlsx"
OUT = base / "GM_RECHECK_2026-08-03.json"

PREV_CONF = "87EF7439A8F19C52B5B948D379D1752B327144CCDBC36758A780A93EC71121FB"
PREV_LIVE = "9F7257A073F37A5822EC8B71882183915E044C768696C5380DC248B98DFCF5D5"


def sha(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest().upper()


def norm(v):
    if v is None:
        return None
    if isinstance(v, ArrayFormula):
        return ("AF", v.ref, v.text)
    return ("V", v)


def main() -> None:
    hc, hl = sha(CONF), sha(LIVE)
    conf_meta = {
        "path": str(CONF),
        "size": CONF.stat().st_size,
        "mtime": datetime.fromtimestamp(CONF.stat().st_mtime).isoformat(timespec="seconds"),
        "sha256": hc,
        "matches_prior_doc_sha": hc == PREV_CONF,
    }
    live_meta = {
        "path": str(LIVE),
        "size": LIVE.stat().st_size,
        "mtime": datetime.fromtimestamp(LIVE.stat().st_mtime).isoformat(timespec="seconds"),
        "sha256": hl,
        "matches_prior_doc_sha": hl == PREV_LIVE,
    }

    print("=== FILE IDENTITY ===")
    print("CONF", conf_meta)
    print("LIVE", live_meta)
    print("byte_identical", CONF.read_bytes() == LIVE.read_bytes())
    print("sha_identical", hc == hl)
    print("size_delta_live_minus_conf", LIVE.stat().st_size - CONF.stat().st_size)
    print("CONF changed vs prior doc SHA:", not conf_meta["matches_prior_doc_sha"])
    print("LIVE changed vs prior doc SHA:", not live_meta["matches_prior_doc_sha"])

    with zipfile.ZipFile(CONF) as zc, zipfile.ZipFile(LIVE) as zl:
        sc = {i.filename: (i.file_size, i.CRC) for i in zc.infolist() if not i.is_dir()}
        sl = {i.filename: (i.file_size, i.CRC) for i in zl.infolist() if not i.is_dir()}
    only_c = sorted(set(sc) - set(sl))
    only_l = sorted(set(sl) - set(sc))
    changed = sorted(k for k in set(sc) & set(sl) if sc[k] != sl[k])
    print()
    print("=== ZIP PARTS ===")
    print("only_confirmed", len(only_c), only_c[:10])
    print("only_live", len(only_l), only_l[:10])
    print("crc_or_size_differ", len(changed))
    for k in changed[:25]:
        print(" ", k, "conf", sc[k], "live", sl[k])

    print()
    print("Loading workbooks...")
    wb_c = load_workbook(CONF, data_only=False)
    wb_l = load_workbook(LIVE, data_only=False)
    wb_cv = load_workbook(CONF, data_only=True)
    wb_lv = load_workbook(LIVE, data_only=True)

    print("sheets", len(wb_c.sheetnames), len(wb_l.sheetnames), "names_equal", wb_c.sheetnames == wb_l.sheetnames)

    diffs = []
    formula_string_diffs = []
    array_formula_diffs = []
    for s in wb_c.sheetnames:
        if s not in wb_l.sheetnames:
            diffs.append({"sheet": s, "cell": "*", "kind": "missing_sheet_live"})
            continue
        wsc, wsl = wb_c[s], wb_l[s]
        coords = set()
        for ws in (wsc, wsl):
            mr = ws.max_row or 1
            mc = ws.max_column or 1
            for row in ws.iter_rows(min_row=1, max_row=mr, max_col=mc):
                for cell in row:
                    if cell.value is not None:
                        coords.add(cell.coordinate)
        for coord in sorted(coords):
            vc, vl = wsc[coord].value, wsl[coord].value
            if norm(vc) == norm(vl):
                continue
            vcv = wb_cv[s][coord].value
            vlv = wb_lv[s][coord].value
            kind = "value"
            if isinstance(vc, ArrayFormula) or isinstance(vl, ArrayFormula):
                kind = "array_formula"
                array_formula_diffs.append(
                    {
                        "sheet": s,
                        "cell": coord,
                        "conf_ref": getattr(vc, "ref", None) if isinstance(vc, ArrayFormula) else None,
                        "live_ref": getattr(vl, "ref", None) if isinstance(vl, ArrayFormula) else None,
                        "conf_text": getattr(vc, "text", None) if isinstance(vc, ArrayFormula) else str(vc)[:100],
                        "live_text": getattr(vl, "text", None) if isinstance(vl, ArrayFormula) else str(vl)[:100],
                    }
                )
            elif (isinstance(vc, str) and vc.startswith("=")) or (isinstance(vl, str) and str(vl).startswith("=")):
                kind = "formula_string"
                formula_string_diffs.append({"sheet": s, "cell": coord, "conf": str(vc)[:200], "live": str(vl)[:200]})
            item = {
                "sheet": s,
                "cell": coord,
                "kind": kind,
                "conf": str(vc)[:160] if vc is not None else None,
                "live": str(vl)[:160] if vl is not None else None,
                "conf_cached": str(vcv)[:100] if vcv is not None else None,
                "live_cached": str(vlv)[:100] if vlv is not None else None,
                "sheet_state": wsc.sheet_state,
            }
            diffs.append(item)

    print()
    print("=== CELL CONTENT DIFFS ===")
    print("total", len(diffs))
    print("formula_string_diffs", len(formula_string_diffs))
    print("array_formula_diffs", len(array_formula_diffs))
    by_sheet: dict[str, list] = {}
    for d in diffs:
        by_sheet.setdefault(d["sheet"], []).append(d)
    for s, items in sorted(by_sheet.items(), key=lambda x: (-len(x[1]), x[0])):
        print(f"--- {s} ({wb_c[s].sheet_state}): {len(items)} ---")
        for d in items[:40]:
            print(
                f"  {d['cell']} [{d['kind']}] conf={d['conf']!r} live={d['live']!r} "
                f"cache_c={d.get('conf_cached')!r} cache_l={d.get('live_cached')!r}"
            )
        if len(items) > 40:
            print(f"  ... +{len(items) - 40} more")

    # Equity Dash focus
    print()
    print("=== EQUITY DASH FOCUS ===")
    for label, wb, wbv in [("CONF", wb_c, wb_cv), ("LIVE", wb_l, wb_lv)]:
        ws, wsv = wb["Equity Dash"], wbv["Equity Dash"]
        print(label, "merges", sorted(str(x) for x in ws.merged_cells.ranges))
        for coord in ["C4", "C5", "C6", "C8", "C9", "C10", "C11", "A4", "A8"]:
            print(f"  {label} {coord} F={ws[coord].value!r} V={wsv[coord].value!r}")
        for coord in ["M4", "N4", "O4", "P4", "Q4", "R4", "O5", "P5", "O9", "R9"]:
            v = ws[coord].value
            if isinstance(v, ArrayFormula):
                print(f"  {label} {coord} AF ref={v.ref} text={(v.text or '')[:90]} cache={wsv[coord].value}")
            else:
                print(f"  {label} {coord} type={type(v).__name__} val={v!r} cache={wsv[coord].value}")

    # named ranges
    def names(wb):
        out = {}
        for name in wb.defined_names:
            dn = wb.defined_names[name]
            key = name if isinstance(name, str) else getattr(dn, "name", str(name))
            out[key] = getattr(dn, "attr_text", str(dn))
        return out

    nc, nl = names(wb_c), names(wb_l)
    named_changed = [k for k in nc if k in nl and nc[k] != nl[k]]
    print()
    print("=== NAMED RANGES ===")
    print("count", len(nc), len(nl), "sets_equal", set(nc) == set(nl), "defs_changed", len(named_changed))

    # classify residual
    residual_kinds = {}
    for d in diffs:
        residual_kinds[d["kind"]] = residual_kinds.get(d["kind"], 0) + 1

    equity_aligned = (
        wb_c["Equity Dash"]["C4"].value == wb_l["Equity Dash"]["C4"].value
        and str(wb_c["Equity Dash"]["C5"].value) == str(wb_l["Equity Dash"]["C5"].value)
        and wb_c["Equity Dash"]["C9"].value == wb_l["Equity Dash"]["C9"].value
        and wb_c["Equity Dash"]["C11"].value == wb_l["Equity Dash"]["C11"].value
        and not any(d["sheet"] == "Equity Dash" for d in diffs)
    )

    # disposition heuristic
    if hc == hl:
        disposition = "A_IDENTICAL"
        disposition_note = "Byte/SHA identical"
    elif len(formula_string_diffs) == 0 and len(array_formula_diffs) == 0 and equity_aligned:
        # remaining are value-only; check if material
        material_value_sheets = sorted({d["sheet"] for d in diffs if d["kind"] == "value"})
        disposition = "B_OR_C_VALUE_RESIDUAL"
        disposition_note = (
            f"No formula/array structural diffs; Equity Dash cell content aligned; "
            f"value residuals on sheets: {material_value_sheets}. "
            f"ZIP/CRC still differ ({len(changed)} parts). "
            f"BOTH files SHA drifted from prior documented baseline."
        )
    else:
        disposition = "B_SUBSTANTIVE"
        disposition_note = "Formula/array or Equity structural residuals remain"

    print()
    print("=== DISPOSITION HEURISTIC ===")
    print("equity_dash_content_aligned", equity_aligned)
    print("disposition", disposition)
    print("note", disposition_note)
    print("prior_CONF_sha", PREV_CONF)
    print("prior_LIVE_sha", PREV_LIVE)
    print("CRITICAL: confirmed GM file on disk no longer matches documented SHA" if hc != PREV_CONF else "CONF matches prior doc")
    print("CRITICAL: live file on disk no longer matches documented SHA" if hl != PREV_LIVE else "LIVE matches prior doc")

    report = {
        "recheck_utc": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "confirmed": conf_meta,
        "live": live_meta,
        "prior_documented": {"confirmed_sha": PREV_CONF, "live_sha": PREV_LIVE},
        "byte_identical": CONF.read_bytes() == LIVE.read_bytes(),
        "zip": {
            "only_confirmed": only_c,
            "only_live": only_l,
            "changed_count": len(changed),
            "changed_sample": changed[:40],
        },
        "sheet_names_equal": wb_c.sheetnames == wb_l.sheetnames,
        "named_ranges": {
            "count_conf": len(nc),
            "count_live": len(nl),
            "sets_equal": set(nc) == set(nl),
            "defs_changed": named_changed,
        },
        "cell_diffs_total": len(diffs),
        "formula_string_diffs": formula_string_diffs,
        "array_formula_diffs": array_formula_diffs,
        "value_diffs": [d for d in diffs if d["kind"] == "value"],
        "equity_dash_content_aligned": equity_aligned,
        "disposition": disposition,
        "disposition_note": disposition_note,
        "residual_kinds": residual_kinds,
    }
    OUT.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
    print()
    print("Wrote", OUT)


if __name__ == "__main__":
    main()
