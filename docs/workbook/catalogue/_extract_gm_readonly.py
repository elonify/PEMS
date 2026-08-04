"""
Read-only extraction from PEMS Golden Master.
Does not write to or modify the xlsx.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

# Authoritative Golden Master = confirmed history snapshot (not live working copy).
GM = Path(
    r"C:\Users\Emmanuel Onwuka\Desktop\PEMS\docs\workbook\Workbook_History"
    r"\Econ_Model_PEMS_confirmed_2026-08-03.xlsx"
)
OUT = Path(r"C:\Users\Emmanuel Onwuka\Desktop\PEMS\docs\workbook\catalogue")
VAL = Path(r"C:\Users\Emmanuel Onwuka\Desktop\PEMS\docs\workbook\Validation_Datasets")
EXPECTED_ACTIVE_SHA = (
    "D07560CA6C1A762716E1927A130E7CA697DB0AB9BDA8E8A33C7A0ACBB6FDBFEA"
)
PREVIOUS_DOCUMENTED_GM_SHA_SUPERSEDED = (
    "87EF7439A8F19C52B5B948D379D1752B327144CCDBC36758A780A93EC71121FB"
)
OUT.mkdir(parents=True, exist_ok=True)
(VAL / "scenarios").mkdir(parents=True, exist_ok=True)
(VAL / "expected_outputs").mkdir(parents=True, exist_ok=True)
(VAL / "regression").mkdir(parents=True, exist_ok=True)


def module_for_sheet(name: str) -> str:
    """Provisional PEMS module area from worksheet name only (not invented business rules)."""
    n = name.lower()
    if any(x in n for x in ("input", "start", "master", "checklist", "model map", "ec_io")):
        return "Input / Control"
    if "fiscal" in n or "pia" in n:
        return "Fiscal Terms"
    if any(x in n for x in ("stoiip", "giip")):
        return "Reservoir"
    if "prod" in n or "production" in n or "block_oil" in n or "block_gas" in n or "oml" in n:
        return "Production"
    if "block_tc" in n or "cap_allow" in n:
        return "Cost / Capital Allowance"
    if "royalt" in n or n == "flgt":
        return "Fiscal / Royalty"
    if "ht_ncf" in n or "cit_ncf" in n or "project_ncf" in n or "equity_ncf" in n or "ncf" in n:
        return "Cash Flow / Tax NCF"
    if "result" in n or "analysis" in n or "dash" in n or "cr econ" in n:
        return "Results / Dashboard"
    if n in ("end", "sheet1"):
        return "Unclassified / Navigation"
    return "Unclassified — needs human review"


def serialize(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float):
        return format(v, ".15g")
    return str(v)


def extract_precedents(formula: str) -> list[str]:
    """Best-effort Excel reference extraction from formula text (not a full parser)."""
    if not formula:
        return []
    f = formula[1:] if formula.startswith("=") else formula
    refs: set[str] = set()
    pattern = re.compile(
        r"(?:(?:'[^']+')|(?:[A-Za-z_][\w.]*))?!\$?[A-Za-z]{1,3}\$?\d+(?::\$?[A-Za-z]{1,3}\$?\d+)?"
    )
    for m in pattern.finditer(f):
        refs.add(m.group(0))
    bare = re.compile(
        r"(?<![A-Za-z0-9_'!])(\$?[A-Za-z]{1,3}\$?\d+(?::\$?[A-Za-z]{1,3}\$?\d+)?)"
    )
    for m in bare.finditer(f):
        refs.add(m.group(1))
    return sorted(refs)[:80]


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


def main() -> None:
    sha = hashlib.sha256(GM.read_bytes()).hexdigest().upper()
    print("SHA256", sha)
    if sha != EXPECTED_ACTIVE_SHA:
        raise SystemExit(
            f"Refusing extract: SHA {sha} != expected active {EXPECTED_ACTIVE_SHA}"
        )
    print("Active Golden Master SHA verified.")
    print("Loading formulas...")
    wb_f = load_workbook(GM, data_only=False, read_only=False, keep_vba=True)
    print("Loading cached values...")
    wb_v = load_workbook(GM, data_only=True, read_only=False)

    sheet_meta: list[dict] = []
    all_rows: list[dict] = []
    formula_rows: list[dict] = []
    value_rows: list[dict] = []
    dv_rows: list[dict] = []
    comment_rows: list[dict] = []

    sheets = wb_f.sheetnames
    print("Sheets", len(sheets))

    for sname in sheets:
        ws = wb_f[sname]
        ws_v = wb_v[sname]
        state = "hidden" if ws.sheet_state and ws.sheet_state != "visible" else "visible"
        mod = module_for_sheet(sname)

        try:
            dvs = list(ws.data_validations.dataValidation) if ws.data_validations else []
        except Exception:
            dvs = []
        for dv in dvs:
            dv_rows.append(
                {
                    "worksheet": sname,
                    "sqref": str(dv.sqref),
                    "type": dv.type,
                    "operator": dv.operator,
                    "formula1": dv.formula1,
                    "formula2": dv.formula2,
                    "allow_blank": dv.allow_blank,
                    "showDropDown": getattr(dv, "showDropDown", None),
                    "module_area": mod,
                }
            )

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        n_form = n_const = n_label = n_nonempty = 0

        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                n_nonempty += 1
                coord = cell.coordinate
                has_f = isinstance(v, str) and v.startswith("=")
                try:
                    cached = ws_v[coord].value
                except Exception:
                    cached = None

                if has_f:
                    n_form += 1
                    formula = v
                    precedents = extract_precedents(formula)
                    entry = {
                        "golden_master_sha256": sha,
                        "workbook_version_label": "Confirmed-2026-08-03",
                        "worksheet": sname,
                        "sheet_state": state,
                        "cell": coord,
                        "row": cell.row,
                        "col": cell.column,
                        "cell_class": "formula",
                        "formula": formula,
                        "cached_value": serialize(cached),
                        "cached_value_type": type(cached).__name__ if cached is not None else "",
                        "number_format": cell.number_format,
                        "precedents_extracted": " | ".join(precedents),
                        "precedent_count": len(precedents),
                        "module_area": mod,
                        "data_type_openpyxl": cell.data_type,
                        "ambiguity_flag": "NO_CACHED_VALUE" if cached is None else "",
                        "notes": (
                            "Formula present but data_only cache empty; full Excel recalc may be required for expected value."
                            if cached is None
                            else "Formula captured from Golden Master; cached value from workbook value cache."
                        ),
                    }
                    formula_rows.append(entry)
                    all_rows.append(entry)
                    if cached is not None:
                        value_rows.append(
                            {
                                "golden_master_sha256": sha,
                                "workbook_version_label": "Confirmed-2026-08-03",
                                "worksheet": sname,
                                "cell": coord,
                                "kind": (
                                    "formula_result_text"
                                    if isinstance(cached, str)
                                    else "formula_result"
                                ),
                                "expected_value": serialize(cached),
                                "value_type": type(cached).__name__,
                                "formula": formula,
                                "number_format": cell.number_format,
                                "module_area": mod,
                                "sheet_state": state,
                                "use_as_numeric_golden": (
                                    "NO"
                                    if isinstance(cached, str)
                                    and str(cached).startswith("#")
                                    else "YES"
                                ),
                            }
                        )
                else:
                    if isinstance(v, str):
                        n_label += 1
                        ctype = "label_or_text"
                        notes = "Text/label literal from Golden Master."
                    else:
                        n_const += 1
                        ctype = "constant_or_input_value"
                        notes = "Numeric/date/bool literal — candidate INPUT or PARAMETER (Excel does not mark input vs constant)."
                    entry = {
                        "golden_master_sha256": sha,
                        "workbook_version_label": "Confirmed-2026-08-03",
                        "worksheet": sname,
                        "sheet_state": state,
                        "cell": coord,
                        "row": cell.row,
                        "col": cell.column,
                        "cell_class": ctype,
                        "formula": "",
                        "cached_value": serialize(v),
                        "cached_value_type": type(v).__name__,
                        "number_format": cell.number_format,
                        "precedents_extracted": "",
                        "precedent_count": 0,
                        "module_area": mod,
                        "data_type_openpyxl": cell.data_type,
                        "ambiguity_flag": (
                            "INPUT_VS_CONSTANT_AMBIGUOUS"
                            if ctype == "constant_or_input_value"
                            else ""
                        ),
                        "notes": notes,
                    }
                    all_rows.append(entry)
                    if ctype == "constant_or_input_value":
                        value_rows.append(
                            {
                                "golden_master_sha256": sha,
                                "workbook_version_label": "Confirmed-2026-08-03",
                                "worksheet": sname,
                                "cell": coord,
                                "kind": "literal_value",
                                "expected_value": serialize(v),
                                "value_type": type(v).__name__,
                                "formula": "",
                                "number_format": cell.number_format,
                                "module_area": mod,
                                "sheet_state": state,
                                "use_as_numeric_golden": "YES",
                            }
                        )

        # comments (openpyxl worksheet comments map if present)
        try:
            comments_map = getattr(ws, "_comments", None) or {}
            if hasattr(ws, "comments") and not comments_map:
                # fallback: iterate known cells with comments via _cells
                comments_map = {}
            for ref, c in (comments_map.items() if hasattr(comments_map, "items") else []):
                comment_rows.append(
                    {
                        "worksheet": sname,
                        "cell": str(ref),
                        "author": getattr(c, "author", "") or "",
                        "text": str(
                            getattr(c, "content", None)
                            or getattr(c, "text", None)
                            or c
                            or ""
                        )[:2000],
                        "module_area": mod,
                    }
                )
        except Exception as ex:
            comment_rows.append(
                {
                    "worksheet": sname,
                    "cell": "",
                    "author": "",
                    "text": f"COMMENT_SCAN_ERROR: {ex}",
                    "module_area": mod,
                }
            )

        sheet_meta.append(
            {
                "worksheet": sname,
                "sheet_state": state,
                "module_area": mod,
                "max_row": max_row,
                "max_col": max_col,
                "dimension": ws.calculate_dimension(),
                "nonempty_cells": n_nonempty,
                "formula_cells": n_form,
                "label_or_text_cells": n_label,
                "constant_or_input_cells": n_const,
                "data_validations": len(dvs),
            }
        )
        print(
            f"  {sname}: nonempty={n_nonempty} formulas={n_form} constants={n_const} labels={n_label}"
        )

    names = []
    try:
        # openpyxl 3.x: DefinedNameDict
        for name in wb_f.defined_names:
            dn = wb_f.defined_names[name]
            names.append(
                {
                    "name": name if isinstance(name, str) else getattr(dn, "name", str(name)),
                    "attr_text": getattr(dn, "attr_text", str(dn)),
                    "hidden": bool(getattr(dn, "hidden", False)),
                }
            )
    except Exception as ex:
        names.append({"name": "DEFINED_NAME_SCAN_ERROR", "attr_text": str(ex), "hidden": False})

    if sheet_meta:
        write_csv(OUT / "sheet_summary.csv", sheet_meta, list(sheet_meta[0].keys()))
    if formula_rows:
        write_csv(OUT / "formula_catalogue.csv", formula_rows, list(formula_rows[0].keys()))
    if all_rows:
        write_csv(
            OUT / "cell_catalogue_all_nonempty.csv", all_rows, list(all_rows[0].keys())
        )
    if dv_rows:
        write_csv(OUT / "data_validations.csv", dv_rows, list(dv_rows[0].keys()))
    else:
        write_csv(
            OUT / "data_validations.csv",
            [],
            [
                "worksheet",
                "sqref",
                "type",
                "operator",
                "formula1",
                "formula2",
                "allow_blank",
                "showDropDown",
                "module_area",
            ],
        )
    if comment_rows:
        write_csv(OUT / "comments.csv", comment_rows, list(comment_rows[0].keys()))
    else:
        write_csv(
            OUT / "comments.csv",
            [],
            ["worksheet", "cell", "author", "text", "module_area"],
        )
    write_csv(OUT / "defined_names.csv", names, ["name", "attr_text", "hidden"])

    formula_expected = [r for r in value_rows if r["kind"].startswith("formula_result")]
    literal_inputs = [r for r in value_rows if r["kind"] == "literal_value"]

    if formula_expected:
        write_csv(
            VAL / "expected_outputs" / "formula_cached_results_all.csv",
            formula_expected,
            list(formula_expected[0].keys()),
        )
    if literal_inputs:
        write_csv(
            VAL / "expected_outputs" / "literal_values_all.csv",
            literal_inputs,
            list(literal_inputs[0].keys()),
        )

    summary = {
        "golden_master": str(GM),
        "golden_master_relative": (
            "docs/workbook/Workbook_History/Econ_Model_PEMS_confirmed_2026-08-03.xlsx"
        ),
        "workbook_version_label": "Confirmed-2026-08-03",
        "sha256": sha,
        "baseline_status": "ACTIVE",
        "previous_documented_gm_sha256_superseded": PREVIOUS_DOCUMENTED_GM_SHA_SUPERSEDED,
        "historical_intake_sha256": "F6A1992F6A3CC27EC587779ADE6CF667B246FB1587296EFD0CD14B47A6783006",
        "extracted_at": datetime.now().isoformat(timespec="seconds"),
        "worksheets_inspected": len(sheets),
        "worksheet_names": sheets,
        "formula_cells": len(formula_rows),
        "formulas_with_cached_value": sum(
            1 for r in formula_rows if r["cached_value"] != ""
        ),
        "formulas_without_cached_value": sum(
            1 for r in formula_rows if r["cached_value"] == ""
        ),
        "nonempty_cells_total": len(all_rows),
        "data_validations": len(dv_rows),
        "comments": len(comment_rows),
        "defined_names": len(names),
        "formula_expected_outputs": len(formula_expected),
        "literal_values_captured": len(literal_inputs),
        "ambiguous_formula_no_cache": sum(
            1 for r in formula_rows if r["ambiguity_flag"] == "NO_CACHED_VALUE"
        ),
        "ambiguous_input_vs_constant": sum(
            1 for r in all_rows if r.get("ambiguity_flag") == "INPUT_VS_CONSTANT_AMBIGUOUS"
        ),
        "by_module_formulas": dict(Counter(r["module_area"] for r in formula_rows)),
        "by_sheet": sheet_meta,
    }
    (OUT / "extraction_summary.json").write_text(
        json.dumps(summary, indent=2), encoding="utf-8"
    )
    print("DONE")
    slim = {k: v for k, v in summary.items() if k != "by_sheet"}
    print(json.dumps(slim, indent=2))


if __name__ == "__main__":
    main()
