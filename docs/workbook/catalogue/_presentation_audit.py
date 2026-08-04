"""
Read-only presentation audit of approved Golden Master.
Does NOT write to or save the xlsx.
"""
from __future__ import annotations

import hashlib
import json
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles.numbers import is_date_format

ROOT = Path(r"C:\Users\Emmanuel Onwuka\Desktop\PEMS")
GM = ROOT / "docs/workbook/Workbook_History/Econ_Model_PEMS_confirmed_2026-08-03.xlsx"
OUT = ROOT / "docs/02_SPECIFICATIONS/presentation/PRESENTATION_AUDIT_EXTRACT.json"
EXPECTED_SHA = "D07560CA6C1A762716E1927A130E7CA697DB0AB9BDA8E8A33C7A0ACBB6FDBFEA"

# Sample cells of known semantic interest + systematic sample
KEY_CELLS = {
    "Ec_IO": [
        "C4",
        "C5",
        "C7",
        "C12",
        "C14",
        "C15",
        "C17",
        "G18",
        "G20",
        "G22",
        "G24",
        "G25",
        "G26",
        "B5",
        "B12",
        "B15",
        "G7",
        "N16",
        "T7",
        "G11",
    ],
    "Equity Dash": ["B3", "B4", "C4", "C5", "C6", "A4", "C8", "C11"],
    "Fiscal Terms_PIA": ["A1", "S16", "T18", "W18", "U30", "T72", "T73"],
    "RESULTS Equity": ["G7", "H7", "J7", "K7", "N7", "N8", "H26", "J18", "C8"],
    "Project_NCF": ["AF3", "AG3", "AH3", "AG58", "AU14", "AU12", "AF5"],
    "FLGT": ["AB3", "W51", "AB51", "AM51"],
    "Production Profile": ["B2", "C3", "C12", "C9"],
    "Block_TC": ["B2", "C2", "E2", "B3", "A4"],
    "Cap_Allow": ["B2", "FI2", "FL2", "FR5"],
    "Royalties": ["B2", "I4", "P3"],
    "CR Econ": ["E3", "G3", "L3", "M3"],
    "Prod_Summary": ["B1", "V47", "Y48", "AF26"],
    "START": ["A1"],
    "Checklist": ["A1", "B1"],
    "Master": ["A1"],
    "END": ["A1"],
}


def rgb(color) -> str | None:
    if color is None:
        return None
    try:
        if getattr(color, "type", None) == "rgb" and color.rgb:
            return str(color.rgb)
        if getattr(color, "theme", None) is not None:
            return f"theme:{color.theme}"
        if getattr(color, "indexed", None) is not None:
            return f"indexed:{color.indexed}"
    except Exception:
        return None
    return None


def cell_style(cell) -> dict:
    f = cell.font
    fill = cell.fill
    align = cell.alignment
    border = cell.border
    prot = cell.protection
    return {
        "value_type": type(cell.value).__name__,
        "value_preview": str(cell.value)[:80] if cell.value is not None else None,
        "number_format": cell.number_format,
        "font_name": f.name,
        "font_size": f.size,
        "font_bold": bool(f.bold),
        "font_italic": bool(f.italic),
        "font_underline": bool(f.underline) if f.underline else False,
        "font_color": rgb(f.color),
        "fill_type": fill.fill_type,
        "fill_fg": rgb(getattr(fill, "fgColor", None)),
        "fill_bg": rgb(getattr(fill, "bgColor", None)),
        "align_h": align.horizontal,
        "align_v": align.vertical,
        "wrap": align.wrap_text,
        "border_l": border.left.style if border.left else None,
        "border_r": border.right.style if border.right else None,
        "border_t": border.top.style if border.top else None,
        "border_b": border.bottom.style if border.bottom else None,
        "locked": prot.locked if prot else None,
        "hidden_formula": prot.hidden if prot else None,
    }


def main() -> None:
    raw = GM.read_bytes()
    sha = hashlib.sha256(raw).hexdigest().upper()
    st = GM.stat()
    meta = {
        "path": str(GM),
        "sha256": sha,
        "sha_matches_expected": sha == EXPECTED_SHA,
        "size": st.st_size,
        "mtime": datetime.fromtimestamp(st.st_mtime).isoformat(timespec="seconds"),
        "audit_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "read_only": True,
        "workbook_saved": False,
    }
    print("SHA", sha, "OK" if sha == EXPECTED_SHA else "MISMATCH")

    wb = load_workbook(GM, data_only=False, read_only=False, keep_vba=True)

    visible = []
    hidden = []
    sheet_info = {}
    nf_counter: Counter = Counter()
    font_counter: Counter = Counter()
    fill_counter: Counter = Counter()
    size_counter: Counter = Counter()
    key_styles = {}
    dv_by_sheet = {}
    cf_by_sheet = {}
    merges_by_sheet = {}
    dims_by_sheet = {}

    for name in wb.sheetnames:
        ws = wb[name]
        state = ws.sheet_state or "visible"
        entry = {
            "name": name,
            "state": state,
            "sheet_format": {
                "default_row_height": ws.sheet_format.defaultRowHeight
                if ws.sheet_format
                else None,
                "default_col_width": ws.sheet_format.defaultColWidth
                if ws.sheet_format
                else None,
            },
            "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
            "sheet_view": None,
            "tab_color": rgb(ws.sheet_properties.tabColor)
            if ws.sheet_properties and ws.sheet_properties.tabColor
            else None,
            "print_title_rows": ws.print_title_rows,
            "print_title_cols": ws.print_title_cols,
            "page_orientation": ws.page_setup.orientation if ws.page_setup else None,
            "paper_size": ws.page_setup.paperSize if ws.page_setup else None,
            "protection": {
                "sheet": bool(ws.protection.sheet) if ws.protection else False,
                "password": bool(ws.protection.password) if ws.protection else False,
                "enable": {
                    "formatCells": getattr(ws.protection, "formatCells", None),
                    "insertRows": getattr(ws.protection, "insertRows", None),
                    "deleteRows": getattr(ws.protection, "deleteRows", None),
                },
            },
            "max_row": ws.max_row,
            "max_col": ws.max_column,
        }
        if state == "visible":
            visible.append(name)
        else:
            hidden.append(name)
        sheet_info[name] = entry

        if state != "visible":
            continue

        # merges
        merges = [str(m) for m in ws.merged_cells.ranges]
        merges_by_sheet[name] = merges[:50]
        if len(merges) > 50:
            merges_by_sheet[name].append(f"... +{len(merges) - 50} more")

        # column widths sample first 20
        dims = {}
        for i, col in enumerate(ws.column_dimensions):
            if i >= 25:
                break
            cd = ws.column_dimensions[col]
            dims[col] = {"width": cd.width, "hidden": cd.hidden}
        # row heights sample
        rh = {}
        for r in range(1, min(15, (ws.max_row or 1) + 1)):
            rd = ws.row_dimensions[r]
            if rd.height is not None or rd.hidden:
                rh[str(r)] = {"height": rd.height, "hidden": rd.hidden}
        dims_by_sheet[name] = {"cols": dims, "rows": rh}

        # data validations
        try:
            dvs = []
            if ws.data_validations:
                for dv in ws.data_validations.dataValidation:
                    dvs.append(
                        {
                            "sqref": str(dv.sqref),
                            "type": dv.type,
                            "operator": dv.operator,
                            "formula1": dv.formula1,
                            "formula2": dv.formula2,
                            "allow_blank": dv.allow_blank,
                            "showErrorMessage": dv.showErrorMessage,
                            "errorTitle": dv.errorTitle,
                            "error": dv.error,
                            "promptTitle": dv.promptTitle,
                            "prompt": dv.prompt,
                        }
                    )
            dv_by_sheet[name] = dvs
        except Exception as ex:
            dv_by_sheet[name] = [{"error": str(ex)}]

        # conditional formatting
        try:
            cfs = []
            for sqref, rules in ws.conditional_formatting._cf_rules.items():
                for rule in rules:
                    cfs.append(
                        {
                            "sqref": str(sqref),
                            "type": getattr(rule, "type", type(rule).__name__),
                            "priority": getattr(rule, "priority", None),
                            "dxf": str(getattr(rule, "dxf", None))[:120],
                            "formula": getattr(rule, "formula", None),
                            "operator": getattr(rule, "operator", None),
                        }
                    )
            cf_by_sheet[name] = cfs[:40]
        except Exception as ex:
            cf_by_sheet[name] = [{"error": str(ex)}]

        # sample styles: key cells + scan nonempty for top formats
        styles = {}
        for coord in KEY_CELLS.get(name, []):
            try:
                styles[coord] = cell_style(ws[coord])
            except Exception as ex:
                styles[coord] = {"error": str(ex)}

        # systematic sample first 80 nonempty-ish cells in used range
        sample_n = 0
        mr = min(ws.max_row or 1, 80)
        mc = min(ws.max_column or 1, 30)
        for row in ws.iter_rows(min_row=1, max_row=mr, max_col=mc):
            for cell in row:
                if cell.value is None and cell.data_type == "n":
                    continue
                # count formats even for empty styled? only with value
                if cell.value is None:
                    continue
                nf_counter[cell.number_format] += 1
                font_counter[cell.font.name or ""] += 1
                size_counter[str(cell.font.size)] += 1
                ft = cell.fill.fill_type or "none"
                fg = rgb(cell.fill.fgColor) or ""
                fill_counter[f"{ft}:{fg}"] += 1
                sample_n += 1
                if sample_n >= 200:
                    break
            if sample_n >= 200:
                break

        key_styles[name] = styles

    # named styles
    named_styles = []
    try:
        for ns in wb._named_styles:
            named_styles.append(str(ns))
    except Exception:
        named_styles = []

    report = {
        "meta": meta,
        "visible_sheets": visible,
        "hidden_sheets": hidden,
        "sheet_info": sheet_info,
        "key_cell_styles": key_styles,
        "number_format_counts_top": nf_counter.most_common(60),
        "font_name_counts": font_counter.most_common(20),
        "font_size_counts": size_counter.most_common(20),
        "fill_counts_top": fill_counter.most_common(40),
        "data_validations": dv_by_sheet,
        "conditional_formatting": {k: v for k, v in cf_by_sheet.items() if v},
        "merges_sample": merges_by_sheet,
        "dimensions_sample": dims_by_sheet,
        "named_styles": named_styles[:50],
        "workbook_protection": {
            "security_lock_structure": getattr(wb.security, "lockStructure", None)
            if hasattr(wb, "security") and wb.security
            else None,
            "security_lock_windows": getattr(wb.security, "lockWindows", None)
            if hasattr(wb, "security") and wb.security
            else None,
        },
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
    print("Wrote", OUT)
    print("visible", len(visible), visible)
    print("fonts", font_counter.most_common(5))
    print("sizes", size_counter.most_common(8))
    print("nf top", nf_counter.most_common(15))
    # post-check file not modified
    sha2 = hashlib.sha256(GM.read_bytes()).hexdigest().upper()
    print("post_sha", sha2, "unchanged", sha2 == sha)


if __name__ == "__main__":
    main()
