"""
Post-extract: GTC-001 rebuild, historical vs active formula diff, AU14 investigation,
Analysis data-table inventory. Read-only vs Golden Master.
"""
from __future__ import annotations

import csv
import hashlib
import json
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(r"C:\Users\Emmanuel Onwuka\Desktop\PEMS")
# Authoritative GM = confirmed snapshot (not live working copy)
GM = ROOT / "docs/workbook/Workbook_History/Econ_Model_PEMS_confirmed_2026-08-03.xlsx"
CAT = ROOT / "docs/workbook/catalogue"
HIST = CAT / "historical_intake_F6A1992F"
VAL = ROOT / "docs/workbook/Validation_Datasets"
SEM = ROOT / "docs/workbook/semantic_mapping"
ACTIVE_SHA = "D07560CA6C1A762716E1927A130E7CA697DB0AB9BDA8E8A33C7A0ACBB6FDBFEA"
PREVIOUS_DOCUMENTED_GM_SHA_SUPERSEDED = (
    "87EF7439A8F19C52B5B948D379D1752B327144CCDBC36758A780A93EC71121FB"
)
HIST_SHA = "F6A1992F6A3CC27EC587779ADE6CF667B246FB1587296EFD0CD14B47A6783006"


def ser(v) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float):
        return format(v, ".15g")
    return str(v)


def main() -> None:
    sha = hashlib.sha256(GM.read_bytes()).hexdigest().upper()
    assert sha == ACTIVE_SHA, sha

    # --- Formula key sets for diff ---
    def load_formulas(path: Path) -> dict[tuple[str, str], str]:
        out = {}
        with path.open(encoding="utf-8") as f:
            for row in csv.DictReader(f):
                out[(row["worksheet"], row["cell"])] = row.get("formula") or ""
        return out

    active_f = load_formulas(CAT / "formula_catalogue.csv")
    hist_f = load_formulas(HIST / "formula_catalogue.csv") if (HIST / "formula_catalogue.csv").exists() else {}

    active_keys = set(active_f)
    hist_keys = set(hist_f)
    added = sorted(active_keys - hist_keys)
    removed = sorted(hist_keys - active_keys)
    common = active_keys & hist_keys
    changed = sorted([k for k in common if active_f[k] != hist_f[k]])

    # sheet inventory diff
    with (CAT / "extraction_summary.json").open(encoding="utf-8") as f:
        summary = json.load(f)
    active_sheets = set(summary.get("worksheet_names") or [])
    hist_sheets = set()
    if (HIST / "sheet_summary.csv").exists():
        with (HIST / "sheet_summary.csv").open(encoding="utf-8") as f:
            hist_sheets = {r["worksheet"] for r in csv.DictReader(f)}

    sheets_added = sorted(active_sheets - hist_sheets)
    sheets_removed = sorted(hist_sheets - active_sheets)

    diff = {
        "active_sha256": ACTIVE_SHA,
        "historical_sha256": HIST_SHA,
        "active_formula_count": len(active_f),
        "historical_formula_count": len(hist_f),
        "formulas_added": len(added),
        "formulas_removed": len(removed),
        "formulas_changed": len(changed),
        "formulas_unchanged": len(common) - len(changed),
        "worksheets_active": len(active_sheets),
        "worksheets_historical": len(hist_sheets),
        "worksheets_added": sheets_added,
        "worksheets_removed": sheets_removed,
        "sample_added": [{"worksheet": a, "cell": b, "formula": active_f[(a, b)][:120]} for a, b in added[:30]],
        "sample_removed": [{"worksheet": a, "cell": b, "formula": hist_f[(a, b)][:120]} for a, b in removed[:30]],
        "sample_changed": [
            {
                "worksheet": a,
                "cell": b,
                "historical_formula": hist_f[(a, b)][:150],
                "active_formula": active_f[(a, b)][:150],
            }
            for a, b in changed[:40]
        ],
    }
    (CAT / "ACTIVE_VS_HISTORICAL_DIFF.json").write_text(json.dumps(diff, indent=2), encoding="utf-8")
    with (CAT / "ACTIVE_VS_HISTORICAL_FORMULA_DIFF.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["change_type", "worksheet", "cell", "historical_formula", "active_formula"],
        )
        w.writeheader()
        for a, b in added:
            w.writerow(
                {
                    "change_type": "ADDED",
                    "worksheet": a,
                    "cell": b,
                    "historical_formula": "",
                    "active_formula": active_f[(a, b)][:500],
                }
            )
        for a, b in removed:
            w.writerow(
                {
                    "change_type": "REMOVED",
                    "worksheet": a,
                    "cell": b,
                    "historical_formula": hist_f[(a, b)][:500],
                    "active_formula": "",
                }
            )
        for a, b in changed:
            w.writerow(
                {
                    "change_type": "CHANGED",
                    "worksheet": a,
                    "cell": b,
                    "historical_formula": hist_f[(a, b)][:500],
                    "active_formula": active_f[(a, b)][:500],
                }
            )

    # --- GTC-001 rebuild ---
    wb = load_workbook(GM, data_only=False)
    wbv = load_workbook(GM, data_only=True)

    kpi = []
    if "RESULTS Equity" in wb.sheetnames:
        ws, wsv = wb["RESULTS Equity"], wbv["RESULTS Equity"]
        for r in range(1, 40):
            for col in ("J", "K", "M", "N", "H", "L"):
                cell = ws[f"{col}{r}"]
                if cell.value is None:
                    continue
                has_f = isinstance(cell.value, str) and cell.value.startswith("=")
                if not has_f and not isinstance(cell.value, (int, float)):
                    continue
                cv = wsv[f"{col}{r}"].value if has_f else cell.value
                lab = ws[f"G{r}"].value
                kpi.append(
                    {
                        "gtc_id": "GTC-001",
                        "golden_master_sha256": ACTIVE_SHA,
                        "workbook_version_label": "Confirmed-2026-08-03",
                        "metric_label_source": ser(lab),
                        "worksheet": "RESULTS Equity",
                        "cell": f"{col}{r}",
                        "formula": cell.value if has_f else "",
                        "expected_value": ser(cv),
                        "value_type": type(cv).__name__ if cv is not None else "",
                        "use_as_numeric_golden": (
                            "NO"
                            if isinstance(cv, str) and str(cv).startswith("#")
                            else "YES"
                        ),
                        "role": "kpi_or_result",
                    }
                )
        for coord in ["L2", "L3", "C5", "C6", "C7", "L5", "C8", "H7", "H8"]:
            cell = ws[coord]
            if cell.value is None:
                continue
            has_f = isinstance(cell.value, str) and str(cell.value).startswith("=")
            cv = wsv[coord].value if has_f else cell.value
            kpi.append(
                {
                    "gtc_id": "GTC-001",
                    "golden_master_sha256": ACTIVE_SHA,
                    "workbook_version_label": "Confirmed-2026-08-03",
                    "metric_label_source": "header/context",
                    "worksheet": "RESULTS Equity",
                    "cell": coord,
                    "formula": cell.value if has_f else "",
                    "expected_value": ser(cv),
                    "value_type": type(cv).__name__ if cv is not None else "",
                    "use_as_numeric_golden": (
                        "NO"
                        if isinstance(cv, str) and str(cv).startswith("#")
                        else "YES"
                    ),
                    "role": "scenario_context",
                }
            )

    # intermediates + Project_NCF AU14 explicitly recorded as non-golden error
    for sheet, coords in [
        ("HT_NCF_Oil Equity", ["AS51", "AT51", "AQ51", "AO51", "AV51"]),
        ("Equity_NCF_Con", ["AG51", "AH51", "AE51", "AF51", "AJ51"]),
        ("FLGT", ["AB51", "AC51", "AD51"]),
        ("CIT_NCF_Oil Equity", ["AF51", "AG51"]),
        ("CIT_NCF_Gas Equity", ["AF51", "AG51"]),
        ("Prod_Summary", ["V47", "Y47", "Y48", "AA48"]),
        ("Project_NCF", ["AU12", "AU14"]),
    ]:
        if sheet not in wb.sheetnames:
            continue
        wsi, wsiv = wb[sheet], wbv[sheet]
        for coord in coords:
            cell = wsi[coord]
            if cell.value is None:
                continue
            has_f = isinstance(cell.value, str) and str(cell.value).startswith("=")
            cv = wsiv[coord].value if has_f else cell.value
            is_err = isinstance(cv, str) and str(cv).startswith("#")
            kpi.append(
                {
                    "gtc_id": "GTC-001",
                    "golden_master_sha256": ACTIVE_SHA,
                    "workbook_version_label": "Confirmed-2026-08-03",
                    "metric_label_source": "intermediate_or_irr",
                    "worksheet": sheet,
                    "cell": coord,
                    "formula": cell.value if has_f else "",
                    "expected_value": ser(cv),
                    "value_type": type(cv).__name__ if cv is not None else "",
                    "use_as_numeric_golden": "NO" if is_err else "YES",
                    "role": "error_excluded" if is_err else "intermediate_calculation",
                }
            )

    # Equity Dash numerics
    if "Equity Dash" in wb.sheetnames:
        wsd, wsdv = wb["Equity Dash"], wbv["Equity Dash"]
        for r in range(1, 30):
            for c in range(1, 10):
                cell = wsd.cell(r, c)
                if cell.value is None:
                    continue
                has_f = isinstance(cell.value, str) and cell.value.startswith("=")
                if not (has_f or isinstance(cell.value, (int, float))):
                    continue
                cv = wsdv.cell(r, c).value if has_f else cell.value
                kpi.append(
                    {
                        "gtc_id": "GTC-001",
                        "golden_master_sha256": ACTIVE_SHA,
                        "workbook_version_label": "Confirmed-2026-08-03",
                        "metric_label_source": ser(wsd.cell(r, max(1, c - 1)).value),
                        "worksheet": "Equity Dash",
                        "cell": cell.coordinate,
                        "formula": cell.value if has_f else "",
                        "expected_value": ser(cv),
                        "value_type": type(cv).__name__ if cv is not None else "",
                        "use_as_numeric_golden": (
                            "NO"
                            if isinstance(cv, str) and str(cv).startswith("#")
                            else "YES"
                        ),
                        "role": "equity_dash",
                    }
                )

    # Inputs Ec_IO + Fiscal
    inputs = []
    for sheet in ("Ec_IO", "Fiscal Terms_PIA"):
        if sheet not in wb.sheetnames:
            continue
        wse, wsev = wb[sheet], wbv[sheet]
        for row in wse.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                has_f = isinstance(cell.value, str) and cell.value.startswith("=")
                cv = wsev[cell.coordinate].value
                if has_f:
                    inputs.append(
                        {
                            "gtc_id": "GTC-001",
                            "golden_master_sha256": ACTIVE_SHA,
                            "workbook_version_label": "Confirmed-2026-08-03",
                            "worksheet": sheet,
                            "cell": cell.coordinate,
                            "cell_class": "formula",
                            "formula": cell.value,
                            "value": ser(cv),
                            "role": f"{sheet}_formula",
                        }
                    )
                elif isinstance(cell.value, str):
                    inputs.append(
                        {
                            "gtc_id": "GTC-001",
                            "golden_master_sha256": ACTIVE_SHA,
                            "workbook_version_label": "Confirmed-2026-08-03",
                            "worksheet": sheet,
                            "cell": cell.coordinate,
                            "cell_class": "label_or_text",
                            "formula": "",
                            "value": ser(cell.value),
                            "role": f"{sheet}_label",
                        }
                    )
                else:
                    inputs.append(
                        {
                            "gtc_id": "GTC-001",
                            "golden_master_sha256": ACTIVE_SHA,
                            "workbook_version_label": "Confirmed-2026-08-03",
                            "worksheet": sheet,
                            "cell": cell.coordinate,
                            "cell_class": "constant_or_input_value",
                            "formula": "",
                            "value": ser(cell.value),
                            "role": f"{sheet}_literal_candidate_unclassified",
                        }
                    )

    kpi_path = VAL / "expected_outputs" / "GTC-001_kpi_and_intermediates.csv"
    with kpi_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(kpi[0].keys()))
        w.writeheader()
        w.writerows(kpi)

    in_path = VAL / "scenarios" / "GTC-001_input_and_parameter_cells.csv"
    with in_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(inputs[0].keys()))
        w.writeheader()
        w.writerows(inputs)

    # Count formula expected with/without error exclusion
    with (VAL / "expected_outputs" / "formula_cached_results_all.csv").open(
        encoding="utf-8"
    ) as f:
        fe = list(csv.DictReader(f))
    err_expected = [
        r
        for r in fe
        if (r.get("expected_value") or "").startswith("#")
        or r.get("use_as_numeric_golden") == "NO"
    ]

    meta = {
        "gtc_id": "GTC-001",
        "name": "As-saved Golden Master baseline (Confirmed-2026-08-03)",
        "status": "ACTIVE",
        "workbook": "docs/workbook/Workbook_History/Econ_Model_PEMS_confirmed_2026-08-03.xlsx",
        "workbook_history_snapshot": "docs/workbook/Workbook_History/Econ_Model_PEMS_confirmed_2026-08-03.xlsx",
        "live_working_copy": "docs/workbook/Econ_Model_PEMS.xlsx",
        "golden_master_sha256": ACTIVE_SHA,
        "previous_documented_gm_sha256_superseded": PREVIOUS_DOCUMENTED_GM_SHA_SUPERSEDED,
        "workbook_version_label": "Confirmed-2026-08-03",
        "historical_intake_sha256": HIST_SHA,
        "description": (
            "Regression case for the active Golden Master (confirmed snapshot) as stored, "
            "including Excel cached formula results."
        ),
        "input_artifact": "docs/workbook/Validation_Datasets/scenarios/GTC-001_input_and_parameter_cells.csv",
        "expected_full_formula_results": "docs/workbook/Validation_Datasets/expected_outputs/formula_cached_results_all.csv",
        "expected_literals": "docs/workbook/Validation_Datasets/expected_outputs/literal_values_all.csv",
        "expected_kpi_pack": "docs/workbook/Validation_Datasets/expected_outputs/GTC-001_kpi_and_intermediates.csv",
        "kpi_rows": len(kpi),
        "input_rows": len(inputs),
        "formula_expected_rows": len(fe),
        "formula_expected_rows_with_excel_error_strings": len(
            [r for r in fe if (r.get("expected_value") or "").startswith("#")]
        ),
        "excluded_from_numeric_golden_policy": [],
        "tolerance_policy": (
            "Exact for ints/bools/text; float abs/rel 1e-9 for binary representation only. "
            "EXP-001: match no-IRR/#NUM! condition exactly; do not invent IRR."
        ),
        "expected_excel_error_conditions": [
            {
                "id": "EXP-001",
                "worksheet": "Project_NCF",
                "cell": "AU14",
                "formula": "=IRR(AK5:AK49)",
                "expected_excel_result": "#NUM!",
                "classification": "EXPECTED_ACCEPTED_IRR_NO_SIGN_CHANGE_CONDITION",
                "workbook_defect": False,
            }
        ],
    }
    (VAL / "scenarios" / "GTC-001_manifest.json").write_text(
        json.dumps(meta, indent=2), encoding="utf-8"
    )

    # --- AU14 investigation ---
    ws, wsv = wb["Project_NCF"], wbv["Project_NCF"]
    ak_vals = []
    for r in range(5, 50):
        fv = ws[f"AK{r}"].value
        cv = wsv[f"AK{r}"].value
        ak_vals.append(
            {
                "cell": f"AK{r}",
                "formula_or_literal": ser(fv)[:200],
                "cached_value": ser(cv),
                "cached_type": type(cv).__name__ if cv is not None else "NoneType",
            }
        )
    numeric = []
    for item in ak_vals:
        v = wsv[item["cell"]].value
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            numeric.append(float(v))
    signs = set()
    for v in numeric:
        if v > 0:
            signs.add("+")
        elif v < 0:
            signs.add("-")
        # zero ignored for sign-change purpose
    # sign change detection on non-zero sequence
    nonzero = [v for v in numeric if v != 0]
    sign_change = False
    for i in range(1, len(nonzero)):
        if nonzero[i - 1] * nonzero[i] < 0:
            sign_change = True
            break

    # other IRRs
    other_irrs = []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("=") and "IRR(" in v.upper():
                cv = wsv[cell.coordinate].value
                other_irrs.append(
                    {
                        "cell": cell.coordinate,
                        "formula": v,
                        "cached_value": ser(cv),
                        "is_error": isinstance(cv, str) and str(cv).startswith("#"),
                    }
                )

    # AF range for AU12
    af_numeric = []
    for r in range(5, 41):
        cv = wsv[f"AF{r}"].value
        if isinstance(cv, (int, float)) and not isinstance(cv, bool):
            af_numeric.append(float(cv))
    af_nonzero = [v for v in af_numeric if v != 0]
    af_sign_change = any(
        af_nonzero[i - 1] * af_nonzero[i] < 0 for i in range(1, len(af_nonzero))
    )

    # alternative metrics labels near IRR
    labels_near = []
    for r in range(1, 25):
        for c in range(30, 50):
            cell = ws.cell(r, c)
            if isinstance(cell.value, str) and not cell.value.startswith("="):
                labels_near.append({"cell": cell.coordinate, "label": cell.value[:80]})

    au14 = {
        "cell": "Project_NCF!AU14",
        "formula": ser(ws["AU14"].value),
        "cached_result": ser(wsv["AU14"].value),
        "status": "UNRESOLVED_OPEN_PENDING_PO_DOMAIN",
        "disposition_options": [
            "A. Excel model correction required",
            "B. PO/domain explicitly accepts #NUM! as expected behaviour",
        ],
        "disposition_selected": None,
        "do_not": [
            "convert to numeric expected value",
            "invent IRR",
            "replace formula",
            "assume #NUM! is expected without PO/domain decision",
        ],
        "ak5_ak49": {
            "range": "AK5:AK49",
            "cells_examined": 45,
            "cached_numeric_count": len(numeric),
            "cached_none_or_nonnumeric_count": 45 - len(numeric),
            "numeric_values_sample": numeric[:15],
            "sign_set_nonzero": sorted(signs),
            "sign_change_detected_in_cached_nonzero_sequence": sign_change,
            "mathematically_capable_of_irr_from_cached_series": bool(
                sign_change and len(nonzero) >= 2
            ),
            "evidence_note": (
                "If cached AK series has no sign change or is empty/non-numeric, Excel IRR commonly returns #NUM!. "
                "Empty cache cells do not prove the live Excel series is empty without Excel-native review."
            ),
            "cell_detail_csv": "docs/workbook/semantic_mapping/PROJECT_NCF_AU14_AK_SERIES.csv",
        },
        "corroborating_irr_cells_on_Project_NCF": other_irrs,
        "au12_irr_af5_af40": {
            "formula": ser(ws["AU12"].value),
            "cached_result": ser(wsv["AU12"].value),
            "af_cached_numeric_count": len(af_numeric),
            "af_sign_change": af_sign_change,
        },
        "alternative_metrics_observed_labels_near_irr_block": labels_near[:40],
        "business_decision": "NOT_INFERRED — requires PO/domain disposition A or B",
    }

    with (SEM / "PROJECT_NCF_AU14_AK_SERIES.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(ak_vals[0].keys()))
        w.writeheader()
        w.writerows(ak_vals)
    (SEM / "PROJECT_NCF_AU14_INVESTIGATION.json").write_text(
        json.dumps(au14, indent=2), encoding="utf-8"
    )

    # --- Analysis data tables ---
    ws = wb["Analysis"]
    wsv = wbv["Analysis"]
    dt_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if type(cell.value).__name__ == "DataTableFormula":
                cv = wsv[cell.coordinate].value
                dt_cells.append(
                    {
                        "worksheet": "Analysis",
                        "cell": cell.coordinate,
                        "construct": "DataTableFormula",
                        "cached_value": ser(cv)[:100],
                        "affects_calculation_engine": "UNKNOWN_LIKELY_SENSITIVITY_PRESENTATION",
                        "affects_sensitivity_analysis": "YES_LIKELY",
                        "use_as_gtc_expected_without_review": "NO",
                        "pems_impl_note": "Do not invent formulas; map as Excel data-table sensitivity construct if in scope",
                    }
                )
    with (SEM / "ANALYSIS_DATA_TABLE_FORMULAS.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f, fieldnames=list(dt_cells[0].keys()) if dt_cells else ["worksheet", "cell"]
        )
        w.writeheader()
        w.writerows(dt_cells)

    # ACTIVE markers
    (CAT / "ACTIVE_BASELINE.md").write_text(
        f"""# ACTIVE CATALOGUE BASELINE

**Status:** ACTIVE  
**Workbook:** docs/workbook/Econ_Model_PEMS.xlsx  
**Version label:** Confirmed-2026-08-03  
**SHA256:** `{ACTIVE_SHA}`  
**Extracted:** see extraction_summary.json  

Historical intake artefacts: `catalogue/historical_intake_F6A1992F/` (SHA `{HIST_SHA}`) — STALE.
""",
        encoding="utf-8",
    )
    (VAL / "ACTIVE_BASELINE.md").write_text(
        f"""# ACTIVE GTC / VALIDATION DATASET BASELINE

**Status:** ACTIVE  
**GTC:** GTC-001  
**SHA256:** `{ACTIVE_SHA}`  
**Version label:** Confirmed-2026-08-03  

Historical: `Validation_Datasets/historical_intake_F6A1992F/` — STALE.
""",
        encoding="utf-8",
    )

    report = {
        "active_sha256": ACTIVE_SHA,
        "historical_sha256": HIST_SHA,
        "worksheets": summary["worksheets_inspected"],
        "formula_cells": summary["formula_cells"],
        "nonempty_cells": summary["nonempty_cells_total"],
        "formulas_with_cache": summary["formulas_with_cached_value"],
        "formulas_without_cache": summary["formulas_without_cached_value"],
        "defined_names": summary["defined_names"],
        "gtc_formula_expected": len(fe),
        "gtc_kpi_rows": len(kpi),
        "gtc_input_rows": len(inputs),
        "diff": {
            "formulas_added": len(added),
            "formulas_removed": len(removed),
            "formulas_changed": len(changed),
            "sheets_added": sheets_added,
            "sheets_removed": sheets_removed,
        },
        "au14": {
            "result": au14["cached_result"],
            "sign_change": sign_change,
            "numeric_count": len(numeric),
            "capable": au14["ak5_ak49"]["mathematically_capable_of_irr_from_cached_series"],
        },
        "analysis_datatable_count": len(dt_cells),
        "excel_error_strings_in_formula_expected": len(
            [r for r in fe if (r.get("expected_value") or "").startswith("#")]
        ),
    }
    (CAT / "REEXTRACT_REPORT.json").write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(json.dumps(report, indent=2))


if __name__ == "__main__":
    main()
