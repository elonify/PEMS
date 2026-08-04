"""
Semantic mapping phase analysis — read-only vs Golden Master + catalogue.
Does not modify Econ_Model_PEMS.xlsx.
Does not invent classifications for ambiguous literals.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(r"C:\Users\Emmanuel Onwuka\Desktop\PEMS")
GM = ROOT / "docs/workbook/Econ_Model_PEMS.xlsx"
CAT = ROOT / "docs/workbook/catalogue"
OUT = ROOT / "docs/workbook/semantic_mapping"
OUT.mkdir(parents=True, exist_ok=True)

SHEET_REF_RE = re.compile(
    r"(?:'([^']+)'|([A-Za-z0-9_ .]+))!"
)
CELL_REF_RE = re.compile(
    r"(?<![A-Za-z0-9_'!])(\$?[A-Za-z]{1,3}\$?\d+(?::\$?[A-Za-z]{1,3}\$?\d+)?)"
)
REF_ERR_RE = re.compile(r"#REF!", re.I)


def domain_for_sheet(name: str) -> str:
    n = name.lower()
    if name in ("Oil Input", "Gas Input", "START", "Checklist", "Model Map", "Master", "Ec_IO"):
        return "M01_Input_Control"
    if name == "Fiscal Terms_PIA":
        return "M02_Fiscal_Terms"
    if name in ("STOIIP", "GIIP"):
        return "M03_Reservoir"
    if name in (
        "Production Profile",
        "Block_Oil Data",
        "Block_Gas Data",
        "OML123_Oil_S1",
        "Prod_Summary",
    ):
        return "M04_Production"
    if name in ("Block_TC", "Block_TC_Gas", "Cap_Allow", "Cap_Allow Gas"):
        return "M05_Cost_CapAllow"
    if name in ("Royalties", "FLGT"):
        return "M06_Royalty_FLGT"
    if "NCF" in name or name in ("HT_NCF", "CIT_NCF"):
        return "M07_Tax_NCF_Cashflow"
    if name in ("RESULTS Equity", "Equity Dash", "CR Econ"):
        return "M08_Results_Economics"
    if name == "Analysis":
        return "M09_Sensitivity_Analysis"
    if name == "YTD Budget APN (2)":
        return "M99_Unclassified_YTD_Budget"
    if name in ("END", "Sheet1"):
        return "M99_Unclassified_Nav"
    return "M99_Unclassified_Other"


def extract_sheet_deps(formula: str) -> set[str]:
    deps = set()
    if not formula:
        return deps
    for m in SHEET_REF_RE.finditer(formula):
        s = m.group(1) or m.group(2)
        if s:
            deps.add(s.strip())
    return deps


def main() -> None:
    sha = hashlib.sha256(GM.read_bytes()).hexdigest().upper()

    # --- Load formula catalogue streaming ---
    sheet_to_sheets: dict[str, Counter] = defaultdict(Counter)
    formula_count_by_sheet: Counter = Counter()
    ref_err_rows: list[dict] = []
    cr_econ: list[dict] = []
    edges: Counter = Counter()  # (src_sheet, dst_sheet) where formula on dst references src
    name_hits: Counter = Counter()

    # defined names
    with (CAT / "defined_names.csv").open(encoding="utf-8") as f:
        defined = list(csv.DictReader(f))
    name_list = [d["name"] for d in defined if d.get("name")]

    print("Scanning formula_catalogue.csv ...")
    with (CAT / "formula_catalogue.csv").open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            ws = row["worksheet"]
            formula_count_by_sheet[ws] += 1
            formula = row.get("formula") or ""
            cached = row.get("cached_value") or ""
            deps = extract_sheet_deps(formula)
            for d in deps:
                sheet_to_sheets[ws][d] += 1
                edges[(d, ws)] += 1  # d feeds ws
            if REF_ERR_RE.search(cached) or REF_ERR_RE.search(formula):
                ref_err_rows.append(
                    {
                        "worksheet": ws,
                        "cell": row["cell"],
                        "location": "cached_value"
                        if REF_ERR_RE.search(cached)
                        else "formula_text",
                        "formula": formula[:500],
                        "cached_value": cached[:200],
                        "domain": domain_for_sheet(ws),
                        "note": "Do not use as Golden expected value if #REF! in result",
                    }
                )
            if ws == "CR Econ":
                cr_econ.append(row)
            # named range usage (token-ish)
            for nm in name_list:
                if len(nm) < 3:
                    continue
                if nm in formula:
                    name_hits[nm] += 1

    # --- Literals register from cell catalogue ---
    print("Building literal classification register ...")
    lit_rows = []
    lit_count = 0
    with (CAT / "cell_catalogue_all_nonempty.csv").open(encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("cell_class") != "constant_or_input_value":
                continue
            lit_count += 1
            ws = row["worksheet"]
            # Evidence-based classification only:
            # - Do NOT mark as input by default
            # - If sheet is pure results/NCF with almost no formulas adjacent known — still UNCLASSIFIED
            # - If under data validation sqref — could be constrained input CANDIDATE only with DV evidence later
            classification = "UNCLASSIFIED_LITERAL"
            evidence = "Literal numeric/date/bool in GM; Excel does not mark input vs constant."
            decision_needed = (
                "Domain/PO or workbook documentation (Model Map, DV labels, UI forms) "
                "to classify as INPUT, PARAMETER, LOOKUP_TABLE_CONSTANT, or HARDCODED_CONSTANT."
            )
            # Weak evidence: Fiscal Terms + Ec_IO often hold parameters — still only CANDIDATE not confirmed input
            if ws in ("Fiscal Terms_PIA", "Ec_IO", "Equity Dash", "Production Profile"):
                candidate = "PARAMETER_OR_INPUT_CANDIDATE"
                evidence += f" Sheet '{ws}' often holds drivers in this model family; NOT confirmed."
            elif ws in ("Oil Input", "Gas Input"):
                candidate = "INPUT_SHEET_LITERAL_CANDIDATE"
                evidence += " Located on hidden *Input sheet; still may be calculated paste or constant."
            elif "NCF" in ws or ws in ("Royalties", "Prod_Summary", "RESULTS Equity"):
                candidate = "LIKELY_HARDCODED_OR_SEED_ON_CALC_SHEET"
                evidence += " On calculation/results sheet; often seeds/zeros/flags — not assumed input."
            else:
                candidate = "UNCLASSIFIED_LITERAL"

            lit_rows.append(
                {
                    "worksheet": ws,
                    "cell": row["cell"],
                    "value": row.get("cached_value", ""),
                    "value_type": row.get("cached_value_type", ""),
                    "number_format": row.get("number_format", ""),
                    "domain": domain_for_sheet(ws),
                    "classification": classification,
                    "candidate_hint_nonbinding": candidate,
                    "evidence": evidence,
                    "decision_required": decision_needed,
                    "gtc_use": "May appear in literal baseline CSV; not auto-treated as scenario input",
                }
            )

    # write literals in chunks summary + full
    lit_path = OUT / "LITERAL_CLASSIFICATION_REGISTER.csv"
    with lit_path.open("w", newline="", encoding="utf-8") as f:
        fields = list(lit_rows[0].keys()) if lit_rows else []
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(lit_rows)

    lit_summary = Counter(r["domain"] for r in lit_rows)
    cand_summary = Counter(r["candidate_hint_nonbinding"] for r in lit_rows)

    # --- CR Econ analysis ---
    print("CR Econ analysis ...")
    cr_no_cache = []
    cr_with_cache = []
    cr_upstream = Counter()
    cr_downstream_refs = Counter()  # sheets that reference CR Econ
    for row in cr_econ:
        formula = row.get("formula") or ""
        deps = extract_sheet_deps(formula)
        for d in deps:
            cr_upstream[d] += 1
        entry = {
            "cell": row["cell"],
            "formula": formula,
            "cached_value": row.get("cached_value", ""),
            "ambiguity_flag": row.get("ambiguity_flag", ""),
            "upstream_sheets": " | ".join(sorted(deps)),
            "expected_value_status": (
                "MISSING_CACHE_REQUIRES_EXCEL_RECALC_OR_MANUAL"
                if not row.get("cached_value")
                else "CACHE_PRESENT"
            ),
            "safe_to_use_as_gtc_expected": "NO"
            if not row.get("cached_value")
            else "YES_IF_CACHE_TRUSTED",
            "notes": (
                "Do not fabricate expected output. Open in Excel and recalculate, or validate manually."
                if not row.get("cached_value")
                else "Cached value available from GM extraction."
            ),
        }
        if not row.get("cached_value"):
            cr_no_cache.append(entry)
        else:
            cr_with_cache.append(entry)

    # who references CR Econ?
    with (CAT / "formula_catalogue.csv").open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            formula = row.get("formula") or ""
            if "CR Econ" in formula or "'CR Econ'" in formula:
                cr_downstream_refs[row["worksheet"]] += 1

    with (OUT / "CR_ECON_FORMULAS_NO_CACHE.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(cr_no_cache[0].keys()) if cr_no_cache else ["cell"])
        w.writeheader()
        w.writerows(cr_no_cache)
    with (OUT / "CR_ECON_FORMULAS_ALL.csv").open("w", newline="", encoding="utf-8") as f:
        all_cr = cr_no_cache + cr_with_cache
        w = csv.DictWriter(f, fieldnames=list(all_cr[0].keys()) if all_cr else ["cell"])
        w.writeheader()
        w.writerows(all_cr)

    # --- #REF! ---
    with (OUT / "HASHREF_OCCURRENCES.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=list(ref_err_rows[0].keys()) if ref_err_rows else ["worksheet", "cell"],
        )
        w.writeheader()
        w.writerows(ref_err_rows)

    # --- Dependency edges ---
    edge_rows = [
        {
            "from_sheet": a,
            "to_sheet": b,
            "reference_count_in_formulas": c,
            "meaning": f"Formulas on '{b}' reference sheet '{a}'",
        }
        for (a, b), c in edges.most_common()
    ]
    with (OUT / "CROSS_SHEET_DEPENDENCY_EDGES.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "from_sheet",
                "to_sheet",
                "reference_count_in_formulas",
                "meaning",
            ],
        )
        w.writeheader()
        w.writerows(edge_rows)

    # adjacency summary per sheet
    dep_summary = []
    for ws, ctr in sorted(sheet_to_sheets.items()):
        dep_summary.append(
            {
                "worksheet": ws,
                "domain": domain_for_sheet(ws),
                "formula_count": formula_count_by_sheet[ws],
                "upstream_sheets": " | ".join(f"{s}({n})" for s, n in ctr.most_common(25)),
                "upstream_sheet_count": len(ctr),
            }
        )
    with (OUT / "SHEET_UPSTREAM_SUMMARY.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(dep_summary[0].keys()))
        w.writeheader()
        w.writerows(dep_summary)

    # named range usage top
    name_use_rows = [
        {"name": n, "formula_occurrences_substr": c}
        for n, c in name_hits.most_common(100)
    ]
    with (OUT / "NAMED_RANGE_USAGE_TOP100.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["name", "formula_occurrences_substr"])
        w.writeheader()
        w.writerows(name_use_rows)

    # --- Charts via openpyxl ---
    print("Charts inventory ...")
    wb = load_workbook(GM, data_only=False, keep_vba=True)
    chart_rows = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        charts = getattr(ws, "_charts", []) or []
        for i, ch in enumerate(charts):
            title = ""
            try:
                if ch.title:
                    title = str(getattr(ch.title, "tx", ch.title) or ch.title)
            except Exception:
                title = str(getattr(ch, "title", "") or "")
            # series
            series_info = []
            try:
                for s in getattr(ch, "series", []) or []:
                    series_info.append(
                        {
                            "val": str(getattr(s, "val", "")),
                            "cat": str(getattr(s, "cat", "")),
                            "title": str(getattr(s, "title", "")),
                        }
                    )
            except Exception as ex:
                series_info.append({"error": str(ex)})
            chart_rows.append(
                {
                    "worksheet": sname,
                    "chart_index": i,
                    "chart_type": type(ch).__name__,
                    "title": title[:200],
                    "anchor": str(getattr(ch, "anchor", ""))[:120],
                    "series_count": len(getattr(ch, "series", []) or []),
                    "series_preview": json.dumps(series_info[:5])[:1000],
                    "affects": "PRESENTATION_LIKELY — verify series formulas; not calc engine unless linked cells are inputs",
                    "domain": domain_for_sheet(sname),
                }
            )

    with (OUT / "CHART_INVENTORY.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=list(chart_rows[0].keys()) if chart_rows else ["worksheet"],
        )
        w.writeheader()
        w.writerows(chart_rows)

    # VBA
    vba_note = {
        "file_extension": GM.suffix,
        "vba_archive_present": bool(getattr(wb, "vba_archive", None)),
        "package_vbaProject": False,
    }
    import zipfile

    with zipfile.ZipFile(GM) as z:
        vba_note["package_vbaProject"] = any(
            "vbaProject" in n for n in z.namelist()
        )
        vba_note["chart_xml_count"] = sum(
            1 for n in z.namelist() if n.startswith("xl/charts/chart") and n.endswith(".xml")
        )

    # Model Map labels (evidence for semantics)
    model_map = []
    if "Model Map" in wb.sheetnames:
        ws = wb["Model Map"]
        for row in ws.iter_rows(max_row=min(120, ws.max_row or 1), max_col=min(20, ws.max_column or 1)):
            vals = [c.value for c in row if c.value is not None]
            if vals:
                model_map.append([str(v)[:80] for v in vals[:12]])

    # Ec_IO key labels for M01
    ecio_labels = []
    if "Ec_IO" in wb.sheetnames:
        ws = wb["Ec_IO"]
        for r in range(1, min(40, (ws.max_row or 1) + 1)):
            for c in range(1, min(20, (ws.max_column or 1) + 1)):
                v = ws.cell(r, c).value
                if isinstance(v, str) and not v.startswith("="):
                    ecio_labels.append({"cell": ws.cell(r, c).coordinate, "label": v[:120]})

    summary = {
        "sha256": sha,
        "literal_register_rows": lit_count,
        "literal_by_domain": dict(lit_summary),
        "literal_candidate_hints_nonbinding": dict(cand_summary),
        "all_literals_classification": "UNCLASSIFIED_LITERAL (none promoted to INPUT without evidence)",
        "cr_econ_formula_total": len(cr_econ),
        "cr_econ_no_cache": len(cr_no_cache),
        "cr_econ_with_cache": len(cr_with_cache),
        "cr_econ_upstream_sheets": dict(cr_upstream),
        "cr_econ_referenced_by_sheets": dict(cr_downstream_refs),
        "hashref_occurrences": len(ref_err_rows),
        "hashref_by_sheet": dict(Counter(r["worksheet"] for r in ref_err_rows)),
        "cross_sheet_edge_types": len(edges),
        "charts_openpyxl": len(chart_rows),
        "charts_by_sheet": dict(Counter(r["worksheet"] for r in chart_rows)),
        "vba": vba_note,
        "formula_count_by_sheet": dict(formula_count_by_sheet),
        "model_map_sample_rows": model_map[:40],
        "ecio_label_sample": ecio_labels[:50],
    }
    (OUT / "SEMANTIC_PHASE_SUMMARY.json").write_text(
        json.dumps(summary, indent=2), encoding="utf-8"
    )
    print(json.dumps({k: summary[k] for k in summary if k not in ("model_map_sample_rows", "ecio_label_sample")}, indent=2))
    print("DONE")


if __name__ == "__main__":
    main()
