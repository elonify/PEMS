"""
Evidence-based reclassification of critical-path UNRESOLVED literals.
Uses Confirmed-2026-08-03 snapshot (authoritative SHA). Does not modify GM.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

EXPECTED = "87EF7439A8F19C52B5B948D379D1752B327144CCDBC36758A780A93EC71121FB"
GM = Path(
    r"C:\Users\Emmanuel Onwuka\Desktop\PEMS\docs\workbook\Workbook_History\Econ_Model_PEMS_confirmed_2026-08-03.xlsx"
)
OUT = Path(r"C:\Users\Emmanuel Onwuka\Desktop\PEMS\docs\workbook\semantic_mapping")
PREV_UNRES = OUT / "CRITICAL_PATH_LITERALS_UNRESOLVED.csv"

MODULE_BY_SHEET = {
    "Ec_IO": "M-CP-01 Parameters/Ec_IO",
    "Equity Dash": "M-CP-01 Parameters/Equity",
    "Fiscal Terms_PIA": "M-CP-01 Fiscal LAW TABLE",
    "Production Profile": "M-CP-02 Production",
    "Block_Oil Data": "M-CP-02 Production",
    "Block_Gas Data": "M-CP-02 Production",
    "Prod_Summary": "M-CP-02 Production",
    "Block_TC": "M-CP-03 Costs",
    "Block_TC_Gas": "M-CP-03 Costs",
    "Cap_Allow": "M-CP-03 Costs",
    "Cap_Allow Gas": "M-CP-03 Costs",
    "Royalties": "M-CP-04 FLGT/Royalty",
    "FLGT": "M-CP-04 FLGT/Royalty",
    "CR Econ": "M-CP-05 CR/NCF",
    "HT_NCF_Oil": "M-CP-05 CR/NCF",
    "CIT_NCF_Oil": "M-CP-05 CR/NCF",
    "CIT_NCF_Gas": "M-CP-05 CR/NCF",
    "Project_NCF": "M-CP-05 CR/NCF",
    "HT_NCF_Oil Equity": "M-CP-05 CR/NCF",
    "CIT_NCF_Oil Equity": "M-CP-05 CR/NCF",
    "CIT_NCF_Gas Equity": "M-CP-05 CR/NCF",
    "Equity_NCF_Con": "M-CP-05 CR/NCF",
    "RESULTS Equity": "M-CP-06 RESULTS",
}

PATH_SHEETS = list(MODULE_BY_SHEET.keys())


def nearby_labels(ws, r, c, span=8):
    labels = []
    # left
    for dc in range(1, span + 1):
        if c - dc < 1:
            break
        v = ws.cell(r, c - dc).value
        if isinstance(v, str) and not v.startswith("=") and v.strip():
            labels.append(("left", ws.cell(r, c - dc).coordinate, v.strip()[:100]))
            break
    # above
    for dr in range(1, span + 1):
        if r - dr < 1:
            break
        v = ws.cell(r - dr, c).value
        if isinstance(v, str) and not v.startswith("=") and v.strip():
            labels.append(("above", ws.cell(r - dr, c).coordinate, v.strip()[:100]))
            break
    # row header far left A-C
    for cc in range(1, min(4, c)):
        v = ws.cell(r, cc).value
        if isinstance(v, str) and not v.startswith("=") and v.strip():
            labels.append(("row", ws.cell(r, cc).coordinate, v.strip()[:100]))
            break
    # top header row 1-4 same col
    for rr in range(1, min(5, r)):
        v = ws.cell(rr, c).value
        if isinstance(v, str) and not v.startswith("=") and v.strip():
            labels.append(("colhdr", ws.cell(rr, c).coordinate, v.strip()[:100]))
    return labels


def classify(ws_name, cell, val, labels, formula_view_neighbors=""):
    lab = " ".join(x[2] for x in labels).lower()
    mod = MODULE_BY_SHEET.get(ws_name, "UNKNOWN")

    # PO closed overrides
    if ws_name == "Equity Dash" and cell == "C4":
        return (
            "INPUT",
            "PO CLOSED: Equity Dash Share Company 1 is independent user INPUT",
            "HIGH",
            "NONE",
            "RESOLVED",
        )
    if ws_name == "Fiscal Terms_PIA":
        return (
            "CONSTANT",
            "PO CLOSED: Fiscal Terms_PIA is LAW TABLE / regulatory constants",
            "HIGH",
            "NONE",
            "RESOLVED",
        )

    # zeros / ones structural
    if val == 0 or val == 0.0:
        return (
            "DEFAULT/STRUCTURAL",
            "Zero literal common structural seed",
            "MEDIUM",
            "NONE",
            "RESOLVED",
        )
    if val == 1 or val == 1.0:
        if any(k in lab for k in ("project", "flag", "index", "factor", "total", "share")):
            return (
                "DEFAULT/STRUCTURAL",
                f"Unity with structural label context: {lab[:80]}",
                "MEDIUM",
                "NONE",
                "RESOLVED",
            )
        return (
            "DEFAULT/STRUCTURAL",
            "Unity literal; treated as structural unless later contradicted",
            "LOW",
            "NONE",
            "RESOLVED",
        )

    if isinstance(val, bool):
        return ("DEFAULT/STRUCTURAL", "Boolean flag", "HIGH", "NONE", "RESOLVED")

    # year
    if isinstance(val, (int, float)) and 1990 <= float(val) <= 2120 and float(val) == int(float(val)):
        return (
            "ASSUMPTION",
            f"Calendar year-like value; labels={lab[:80]}",
            "MEDIUM",
            "NONE",
            "RESOLVED",
        )

    # small integers count/index
    if isinstance(val, int) and abs(val) <= 20 and val not in (0, 1):
        if any(k in lab for k in ("year", "period", "duration", "payment", "tier", "stage")):
            return (
                "ASSUMPTION",
                f"Small integer with schedule/stage label: {lab[:80]}",
                "MEDIUM",
                "NONE",
                "RESOLVED",
            )
        return (
            "DEFAULT/STRUCTURAL",
            f"Small integer likely index/count: {lab[:80]}",
            "LOW",
            "NONE",
            "RESOLVED",
        )

    # fractions 0-1
    if isinstance(val, float) and 0 < abs(val) <= 1:
        if any(
            k in lab
            for k in (
                "rate",
                "royalt",
                "tax",
                "discount",
                "share",
                "equity",
                "interest",
                "%",
                "allowance",
                "split",
                "limit",
            )
        ):
            # On non-fiscal-law sheets these are case assumptions or coefficients
            if ws_name == "Fiscal Terms_PIA":
                return ("CONSTANT", "Law table rate", "HIGH", "NONE", "RESOLVED")
            if any(k in lab for k in ("share", "equity")) and ws_name == "Equity Dash":
                # other equity fractions if any
                return (
                    "INPUT",
                    f"Share-like on Equity Dash; labels={lab[:80]}",
                    "MEDIUM",
                    "NONE",
                    "RESOLVED",
                )
            return (
                "ASSUMPTION",
                f"Rate/fraction with economic label: {lab[:80]}",
                "MEDIUM",
                "NONE",
                "RESOLVED",
            )
        return (
            "FORMULA COEFFICIENT",
            f"Unit interval without strong economic label: {lab[:80]}",
            "LOW",
            "NONE",
            "RESOLVED",
        )

    # percent-like 1-100
    if isinstance(val, (int, float)) and 1 < abs(float(val)) <= 100:
        if any(k in lab for k in ("%", "percent", "rate", "royalt", "tax", "duration", "year")):
            if ws_name == "Fiscal Terms_PIA":
                return ("CONSTANT", "Law table percent/threshold", "HIGH", "NONE", "RESOLVED")
            return (
                "ASSUMPTION",
                f"Percent/threshold-like: {lab[:80]}",
                "MEDIUM",
                "NONE",
                "RESOLVED",
            )
        if abs(float(val) - int(float(val))) < 1e-12 and int(float(val)) in range(2, 101):
            return (
                "DEFAULT/STRUCTURAL",
                f"Integer 2-100 without rate label: {lab[:80]}",
                "LOW",
                "NONE",
                "RESOLVED",
            )

    # large magnitudes
    if isinstance(val, (int, float)) and abs(float(val)) > 100:
        if any(
            k in lab
            for k in (
                "price",
                "cost",
                "capex",
                "opex",
                "revenue",
                "loan",
                "amount",
                "budget",
                "$",
                "mm",
                "production",
            )
        ):
            return (
                "ASSUMPTION",
                f"Economic magnitude with label: {lab[:80]}",
                "MEDIUM",
                "NONE",
                "RESOLVED",
            )
        return (
            "UNRESOLVED",
            f"Large magnitude without clear label; need domain review: {lab[:80]}",
            "LOW",
            "Domain review of cell purpose / row meaning",
            "UNRESOLVED",
        )

    # residual floats
    if isinstance(val, float):
        return (
            "UNRESOLVED",
            f"Float without sufficient context: {lab[:80]}",
            "LOW",
            "Inspect row/column business meaning in workbook",
            "UNRESOLVED",
        )

    return (
        "UNRESOLVED",
        f"No rule matched: {lab[:80]}",
        "LOW",
        "Manual domain classification required",
        "UNRESOLVED",
    )


def main():
    sha = hashlib.sha256(GM.read_bytes()).hexdigest().upper()
    assert sha == EXPECTED, sha
    wb = load_workbook(GM, data_only=False)

    # collect all critical-path numeric literals (visible)
    all_lits = []
    for sname in PATH_SHEETS:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        if ws.sheet_state == "hidden":
            continue
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None or isinstance(v, str):
                    continue
                if not isinstance(v, (int, float, bool)):
                    continue
                labels = nearby_labels(ws, cell.row, cell.column)
                cls, evidence, conf, decision, status = classify(
                    sname, cell.coordinate, v, labels
                )
                # impact
                impact = "Critical path parameter surface"
                if sname in ("RESULTS Equity",):
                    impact = "Results display/params — low if structural"
                if cls == "INPUT":
                    impact = "Required case input for equity/econ path"
                if sname == "Fiscal Terms_PIA":
                    impact = "Regulatory rate/threshold for fiscal calcs"
                all_lits.append(
                    {
                        "golden_master_sha256": sha,
                        "worksheet": sname,
                        "cell": cell.coordinate,
                        "literal_value": format(v, ".15g")
                        if isinstance(v, float)
                        else repr(v),
                        "formula_containing_literal": "",  # literal cell itself
                        "row_column_labels": " | ".join(f"{a}:{b}:{c}" for a, b, c in labels)[
                            :250
                        ],
                        "surrounding_context": " | ".join(x[2] for x in labels)[:200],
                        "critical_path_module": MODULE_BY_SHEET[sname],
                        "classification": cls,
                        "proposed_classification": cls,
                        "evidence": evidence[:300],
                        "confidence": conf,
                        "impact": impact,
                        "decision_required": decision,
                        "status": status,
                    }
                )

    counts = Counter(r["classification"] for r in all_lits)
    status_counts = Counter(r["status"] for r in all_lits)
    unres = [r for r in all_lits if r["status"] == "UNRESOLVED"]
    resolved = [r for r in all_lits if r["status"] == "RESOLVED"]

    # full register
    reg = OUT / "CRITICAL_PATH_LITERAL_REGISTER_FULL.csv"
    with reg.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(all_lits[0].keys()))
        w.writeheader()
        w.writerows(all_lits)

    unres_path = OUT / "CRITICAL_PATH_LITERALS_UNRESOLVED.csv"
    fields = list(unres[0].keys()) if unres else list(all_lits[0].keys())
    with unres_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(unres)

    # markdown table of remaining unresolved (cap if huge)
    md_lines = [
        "# Critical-Path Unresolved Literals — Exact Location Register",
        "",
        f"**GM SHA:** `{sha}`  ",
        f"**Total critical-path numeric literals:** {len(all_lits)}  ",
        f"**Resolved this pass:** {len(resolved)}  ",
        f"**Still UNRESOLVED:** {len(unres)}  ",
        "",
        "| Sheet | Cell | Value | Context | Module | Proposed Classification | Status | Decision required |",
        "|-------|------|------:|---------|--------|-------------------------|--------|-------------------|",
    ]
    for r in unres:
        ctx = (r["surrounding_context"] or "").replace("|", "/")[:60]
        md_lines.append(
            f"| {r['worksheet']} | {r['cell']} | {r['literal_value']} | {ctx} | {r['critical_path_module']} | {r['proposed_classification']} | UNRESOLVED | {(r['decision_required'] or '')[:80]} |"
        )
    (OUT / "CRITICAL_PATH_UNRESOLVED_LITERALS.md").write_text(
        "\n".join(md_lines) + "\n", encoding="utf-8"
    )

    summary = {
        "sha256": sha,
        "total_critical_path_numeric_literals": len(all_lits),
        "resolved": len(resolved),
        "unresolved": len(unres),
        "by_classification": dict(counts),
        "by_status": dict(status_counts),
        "register_full": str(reg),
        "register_unresolved": str(unres_path),
        "unresolved_markdown": str(OUT / "CRITICAL_PATH_UNRESOLVED_LITERALS.md"),
    }
    (OUT / "CRITICAL_PATH_LITERAL_RESOLUTION_SUMMARY.json").write_text(
        json.dumps(summary, indent=2), encoding="utf-8"
    )
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
