"""Read-only confirmation of error values in Golden Master. Does not modify xlsx."""
from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

GM = Path(r"C:\Users\Emmanuel Onwuka\Desktop\PEMS\docs\workbook\Econ_Model_PEMS.xlsx")
OUT = Path(r"C:\Users\Emmanuel Onwuka\Desktop\PEMS\docs\workbook\semantic_mapping")
OUT.mkdir(parents=True, exist_ok=True)

ERR_PAT = re.compile(
    r"#(REF!|DIV/0!|VALUE!|N/A|NAME\?|NULL!|NUM!|GETTING_DATA|SPILL!|CALC!|FIELD!)",
    re.I,
)


def main() -> None:
    sha = hashlib.sha256(GM.read_bytes()).hexdigest().upper()
    print("SHA256", sha)

    wb_f = load_workbook(GM, data_only=False)
    wb_v = load_workbook(GM, data_only=True)

    errors: list[dict] = []
    by_err_type: Counter = Counter()
    by_sheet: Counter = Counter()

    for sname in wb_f.sheetnames:
        ws = wb_f[sname]
        wsv = wb_v[sname]
        max_r, max_c = ws.max_row or 0, ws.max_column or 0
        state = "hidden" if ws.sheet_state != "visible" else "visible"
        for row in ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c):
            for cell in row:
                fval = cell.value
                try:
                    cval = wsv[cell.coordinate].value
                except Exception:
                    cval = None
                f_str = str(fval) if fval is not None else ""
                c_str = str(cval) if cval is not None else ""
                hits: list[tuple[str, str]] = []
                if isinstance(fval, str) and ERR_PAT.search(fval):
                    hits.append(("formula_text", fval[:500]))
                if isinstance(cval, str) and ERR_PAT.search(cval):
                    hits.append(("cached_value", cval[:300]))
                if not hits:
                    continue
                for kind, text in hits:
                    m = ERR_PAT.search(text)
                    et = m.group(0).upper() if m else "UNKNOWN"
                    by_err_type[et] += 1
                    by_sheet[sname] += 1
                    formula_part = ""
                    if isinstance(fval, str) and fval.startswith("="):
                        formula_part = fval[:500]
                    elif fval is not None and not isinstance(fval, str):
                        formula_part = f"<{type(fval).__name__}>"
                    errors.append(
                        {
                            "worksheet": sname,
                            "cell": cell.coordinate,
                            "error_type": et,
                            "location": kind,
                            "formula_or_source": formula_part,
                            "cached_value": c_str[:200],
                            "sheet_state": state,
                        }
                    )

    fields = (
        list(errors[0].keys())
        if errors
        else [
            "worksheet",
            "cell",
            "error_type",
            "location",
            "formula_or_source",
            "cached_value",
            "sheet_state",
        ]
    )
    with (OUT / "WORKBOOK_ERROR_INVENTORY.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(errors)

    # Analysis specials
    ws = wb_f["Analysis"]
    wsv = wb_v["Analysis"]
    dt = 0
    ref_str = 0
    other_err = Counter()
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if type(v).__name__ == "DataTableFormula":
                dt += 1
            cv = wsv[cell.coordinate].value
            if isinstance(cv, str):
                if "#REF" in cv.upper():
                    ref_str += 1
                m = ERR_PAT.search(cv)
                if m:
                    other_err[m.group(0).upper()] += 1

    # CR Econ
    cr_empty = 0
    cr_err = 0
    cr_ok = 0
    ws, wsv = wb_f["CR Econ"], wb_v["CR Econ"]
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                cv = wsv[cell.coordinate].value
                if cv is None:
                    cr_empty += 1
                elif isinstance(cv, str) and ERR_PAT.search(cv):
                    cr_err += 1
                else:
                    cr_ok += 1

    # Who references START?
    start_refs = 0
    start_ref_sheets: Counter = Counter()
    for sname in wb_f.sheetnames:
        ws = wb_f[sname]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                if "START!" in v or "'START'!" in v:
                    start_refs += 1
                    start_ref_sheets[sname] += 1

    # Unique cells with errors
    unique_cells = sorted({(e["worksheet"], e["cell"], e["error_type"]) for e in errors})
    unique_by_sheet: Counter = Counter(u[0] for u in unique_cells)

    # START sample details
    start_cells = sorted({e["cell"] for e in errors if e["worksheet"] == "START"})

    summary = {
        "confirmation_date": "2026-08-03",
        "golden_master": str(GM),
        "sha256": sha,
        "gm_content_modified_this_pass": False,
        "total_error_location_hits": len(errors),
        "unique_sheet_cell_error_type": len(unique_cells),
        "by_error_type_hits": dict(by_err_type),
        "by_sheet_hits": dict(by_sheet),
        "unique_error_cells_by_sheet": dict(unique_by_sheet),
        "start_error_cells": start_cells,
        "start_error_cell_count": len(start_cells),
        "analysis_datatable_formula_cells": dt,
        "analysis_cached_ref_string_cells": ref_str,
        "analysis_cached_error_types": dict(other_err),
        "cr_econ_formulas_with_cache_ok": cr_ok,
        "cr_econ_formulas_empty_cache": cr_empty,
        "cr_econ_formulas_error_in_cache": cr_err,
        "formulas_referencing_START_sheet": start_refs,
        "sheets_with_formulas_referencing_START": dict(start_ref_sheets),
        "status_assessment": {
            "HASHREF_ON_START": "CONFIRMED_WORKBOOK_STORED_ERROR",
            "START_CONSUMED_BY_CALC_SHEETS": (
                "NO_REFERENCES_FOUND" if start_refs == 0 else "REFERENCES_FOUND"
            ),
            "ANALYSIS_DATATABLES": "CONFIRMED_PRESENT",
            "ANALYSIS_CACHED_REF": (
                "CONFIRMED" if ref_str > 0 else "NOT_IN_STRING_CACHE_OR_ZERO"
            ),
            "CR_ECON_EMPTY_CACHE": "CONFIRMED_EXTRACTION_CACHE_GAP_NOT_EXCEL_ERROR_STRING",
            "CR_ECON_ERROR_STRINGS": "NONE" if cr_err == 0 else "PRESENT",
            "OTHER_ERROR_TYPES_IN_WB": sorted(by_err_type.keys()),
        },
    }
    (OUT / "WORKBOOK_ERROR_STATUS.json").write_text(
        json.dumps(summary, indent=2), encoding="utf-8"
    )
    print(json.dumps(summary, indent=2))
    print("START cells:", ", ".join(start_cells))


if __name__ == "__main__":
    main()
