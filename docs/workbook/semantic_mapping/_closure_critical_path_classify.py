"""Critical-path visible literal classification against Confirmed-2026-08-03 snapshot."""
from __future__ import annotations

import csv
import hashlib
import json
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

EXPECTED = "87EF7439A8F19C52B5B948D379D1752B327144CCDBC36758A780A93EC71121FB"
GM = Path(
    r"C:\Users\Emmanuel Onwuka\Desktop\PEMS\docs\workbook\Workbook_History\Econ_Model_PEMS_confirmed_2026-08-03.xlsx"
)
OUT = Path(r"C:\Users\Emmanuel Onwuka\Desktop\PEMS\docs\workbook\semantic_mapping")
ACTIVE_PATH = Path(r"C:\Users\Emmanuel Onwuka\Desktop\PEMS\docs\workbook\Econ_Model_PEMS.xlsx")

PATH_SHEETS = [
    "Ec_IO",
    "Fiscal Terms_PIA",
    "Equity Dash",
    "Production Profile",
    "Block_Oil Data",
    "Block_Gas Data",
    "Prod_Summary",
    "Block_TC",
    "Block_TC_Gas",
    "Cap_Allow",
    "Cap_Allow Gas",
    "Royalties",
    "FLGT",
    "CR Econ",
    "HT_NCF_Oil",
    "CIT_NCF_Oil",
    "CIT_NCF_Gas",
    "Project_NCF",
    "HT_NCF_Oil Equity",
    "CIT_NCF_Oil Equity",
    "CIT_NCF_Gas Equity",
    "Equity_NCF_Con",
    "RESULTS Equity",
]


def nearby_label(ws, r, c, max_scan=6):
    labels = []
    for dc in range(1, max_scan + 1):
        if c - dc >= 1:
            v = ws.cell(r, c - dc).value
            if isinstance(v, str) and not v.startswith("=") and v.strip():
                labels.append(v.strip()[:80])
                break
    for dr in range(1, max_scan + 1):
        if r - dr >= 1:
            v = ws.cell(r - dr, c).value
            if isinstance(v, str) and not v.startswith("=") and v.strip():
                labels.append(v.strip()[:80])
                break
    return labels


def classify(val, labels, has_dv):
    lab = " ".join(labels).lower()
    if val == 0 or val == 0.0:
        return (
            "DEFAULT_STRUCTURAL_VALUE",
            "Zero literal; not INPUT without further evidence",
            "EXTRACTED",
        )
    if isinstance(val, bool):
        return ("DEFAULT_STRUCTURAL_VALUE", "Boolean literal", "EXTRACTED")
    if isinstance(val, (int, float)) and 1990 <= float(val) <= 2100 and float(val) == int(float(val)):
        depth = "UNDERSTOOD" if lab else "EXTRACTED"
        return ("ASSUMPTION", f"Year-like integer; labels={lab[:80]}", depth)
    if isinstance(val, float) and 0 < val <= 1:
        if has_dv or any(
            k in lab for k in ("rate", "discount", "share", "equity", "interest", "%", "royalt")
        ):
            if has_dv and any(k in lab for k in ("share", "equity", "discount", "input")):
                return (
                    "INPUT",
                    f"DV+label suggests driver; labels={lab[:80]}",
                    "UNDERSTOOD",
                )
            return ("ASSUMPTION", f"Fractional with label/DV; labels={lab[:80]}", "UNDERSTOOD")
        return ("UNRESOLVED", "Fractional literal without strong input evidence", "EXTRACTED")
    if isinstance(val, (int, float)) and 1 < float(val) <= 100:
        if any(k in lab for k in ("%", "percent", "rate", "royalt", "tax", "discount")):
            return ("ASSUMPTION", f"Percent-like with label; labels={lab[:80]}", "UNDERSTOOD")
        return ("UNRESOLVED", "Numeric 1-100 role unclear", "EXTRACTED")
    if isinstance(val, (int, float)) and abs(float(val)) >= 1:
        if any(
            k in lab
            for k in ("price", "cost", "capex", "opex", "$", "mm", "revenue", "production")
        ):
            return ("ASSUMPTION", f"Economic magnitude with label; labels={lab[:80]}", "UNDERSTOOD")
        return ("UNRESOLVED", "Numeric magnitude; role unclear without further evidence", "EXTRACTED")
    return ("UNRESOLVED", "No evidence-based class", "EXTRACTED")


def main():
    sha = hashlib.sha256(GM.read_bytes()).hexdigest().upper()
    if sha != EXPECTED:
        raise SystemExit(f"Confirmed snapshot hash mismatch {sha}")
    path_sha = hashlib.sha256(ACTIVE_PATH.read_bytes()).hexdigest().upper()
    integrity = {
        "confirmed_snapshot_sha256": sha,
        "active_path": str(ACTIVE_PATH),
        "active_path_sha256": path_sha,
        "path_matches_confirmed": path_sha == sha,
        "note": (
            "Classification uses confirmed snapshot. If active path differs, restore or re-intake before implementation."
        ),
    }

    wb = load_workbook(GM, data_only=False)
    dv_map = set()
    for sname in PATH_SHEETS:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        try:
            dvs = list(ws.data_validations.dataValidation) if ws.data_validations else []
        except Exception:
            dvs = []
        for dv in dvs:
            for part in str(dv.sqref).replace(",", " ").split():
                dv_map.add((sname, part.upper()))

    rows = []
    counts = Counter()
    for sname in PATH_SHEETS:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        if ws.sheet_state == "hidden":
            continue
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None or isinstance(v, str):
                    continue
                if not isinstance(v, (int, float, bool)):
                    continue
                labels = nearby_label(ws, cell.row, cell.column)
                has_dv = any(
                    cell.coordinate.upper() in ref or ref.startswith(cell.coordinate[0])
                    for sn, ref in dv_map
                    if sn == sname
                )
                # stricter DV check
                has_dv = False
                for sn, ref in dv_map:
                    if sn == sname and cell.coordinate.upper() in ref.replace("$", ""):
                        has_dv = True
                cls, evidence, depth = classify(v, labels, has_dv)
                counts[cls] += 1
                rows.append(
                    {
                        "golden_master_sha256": sha,
                        "worksheet": sname,
                        "cell": cell.coordinate,
                        "value": format(v, ".15g") if isinstance(v, float) else repr(v),
                        "classification": cls,
                        "understanding": depth,
                        "evidence": evidence[:220],
                        "nearby_labels": " | ".join(labels)[:200],
                        "has_data_validation_hint": has_dv,
                        "critical_path": True,
                        "scope": "VISIBLE_CRITICAL_PATH",
                    }
                )

    # Equity Dash special: find share-like cells
    equity_notes = []
    if "Equity Dash" in wb.sheetnames:
        ws = wb["Equity Dash"]
        for row in ws.iter_rows(max_row=40, max_col=20):
            for cell in row:
                if cell.value is None:
                    continue
                equity_notes.append(
                    {
                        "cell": cell.coordinate,
                        "value": str(cell.value)[:120],
                        "type": type(cell.value).__name__,
                    }
                )

    out = OUT / "CRITICAL_PATH_LITERAL_CLASSIFICATION.csv"
    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    unres = [r for r in rows if r["classification"] == "UNRESOLVED"]
    with (OUT / "CRITICAL_PATH_LITERALS_UNRESOLVED.csv").open(
        "w", newline="", encoding="utf-8"
    ) as f:
        w = csv.DictWriter(f, fieldnames=list(unres[0].keys()) if unres else ["worksheet"])
        w.writeheader()
        w.writerows(unres)

    summary = {
        "source_file": str(GM),
        "sha256": sha,
        "active_path_integrity": integrity,
        "critical_path_literals_classified": len(rows),
        "by_classification": dict(counts),
        "unresolved_count": len(unres),
        "input_count": counts.get("INPUT", 0),
        "assumption_count": counts.get("ASSUMPTION", 0),
        "default_structural_count": counts.get("DEFAULT_STRUCTURAL_VALUE", 0),
        "note": "Never INPUT solely because hard-coded. INPUT only with strong DV+label evidence.",
        "equity_dash_cells_sample": equity_notes[:40],
    }
    (OUT / "CRITICAL_PATH_LITERAL_SUMMARY.json").write_text(
        json.dumps(summary, indent=2), encoding="utf-8"
    )
    print(json.dumps(summary, indent=2)[:4000])


if __name__ == "__main__":
    main()
